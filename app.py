"""
Grist Custom Forms - Flask API Server

Proxies requests to Grist API with authentication.
Supports multiple forms via form_id in URL path.

Environment variables:
  GRIST_API_KEY - Default API key (or GRIST_API_KEY_<FORM_ID> per form)
  GRIST_DOC_<FORM_ID> - Document ID for each form
  GRIST_TABLE_<FORM_ID> - Table ID for each form (defaults to "Reponses")
  GRIST_BASE_URL - Grist instance URL (defaults to grist.numerique.gouv.fr)
"""

import os
import re
import json
import time
import csv
import secrets
from html import escape
from collections import Counter, defaultdict
from datetime import datetime
from io import BytesIO, StringIO
from functools import wraps
from pathlib import Path
from urllib.parse import urljoin
from itsdangerous import BadSignature, URLSafeSerializer
from dotenv import load_dotenv
import requests
from flask import Flask, request, jsonify, redirect, send_file, send_from_directory, Response
from werkzeug.middleware.dispatcher import DispatcherMiddleware

load_dotenv()

BASE_DIR = Path(__file__).parent
FORMS_DIR = BASE_DIR / 'forms'
ASSETS_DIR = BASE_DIR / 'assets'
DOCS_DIR = BASE_DIR / 'docs'

app = Flask(__name__)

GRIST_BASE_URL = os.environ.get('GRIST_BASE_URL', 'https://grist.numerique.gouv.fr').rstrip('/')
_TABLE_COLUMNS_CACHE: dict[tuple[str, str], dict[str, object]] = {}
_TABLE_COLUMNS_CACHE_TTL_SECONDS = 60
EURES_CANDIDATS_TABLE = 'Candidats'
EURES_BESOINS_TABLE = 'Besoins_Employeurs'
EURES_MATCHINGS_TABLE = 'Matchings'
EURES_INVITATIONS_TABLE_DEFAULT = 'Invitations'
EURES_STATS_TABLE_DEFAULT = 'Pilotage_EURES_Mensuel'
EURES_MATCHING_FIELDS = {
    'besoin_id',
    'candidat_id',
    'score',
    'score_metier',
    'score_langues',
    'score_mobilite',
    'score_disponibilite',
    'score_salaire',
    'statut',
    'raisons',
    'points_faibles',
    'date_calcul',
}
EURES_MATCHING_ADMIN_FIELDS = {
    'admin_status',
    'admin_decision_at',
    'admin_decision_by',
    'admin_decision_note',
    'workflow_status',
    'workflow_status_updated_at',
    'workflow_status_updated_by',
    'sent_to_employer_at',
    'sent_to_employer_by',
    'employer_response',
    'employer_response_at',
    'employer_response_source',
    'mise_en_relation_at',
    'mise_en_relation_by',
    'embauche_confirmee_at',
    'embauche_confirmee_by',
}
EURES_INVITATION_FIELDS = {
    'role',
    'email',
    'first_name',
    'last_name',
    'company_name',
    'language',
    'source',
    'external_ref',
    'invite_token',
    'invite_link',
    'invitation_status',
    'invitation_status_updated_at',
    'invitation_status_updated_by',
    'import_batch_id',
    'imported_at',
    'imported_by',
    'sent_at',
    'sent_by',
    'brevo_message_id',
    'answered_at',
    'linked_form_role',
    'linked_record_id',
    'linked_record_key',
    'matching_status',
    'matching_status_updated_at',
    'matching_status_updated_by',
    'notes',
}
EURES_INVITATION_ALLOWED_ROLES = {'candidate', 'employer'}
EURES_INVITATION_ALLOWED_STATUSES = {
    'invitation_a_envoyer',
    'invitation_envoyee',
    'questionnaire_recu',
    'rapprochement_a_confirmer',
    'rapprochee',
    'matching_calcule',
    'traitee',
    'erreur_envoi',
    'desactivee',
}
EURES_INVITATION_SENDABLE_STATUSES = {'invitation_a_envoyer', 'erreur_envoi'}
EURES_SECTOR_CANONICAL_MAP = {
    'vente': 'vente',
    'commerce': 'vente',
    'sales': 'vente',
    'retail': 'vente',
    'nettoyage': 'nettoyage',
    'entretien': 'nettoyage',
    'cleaning': 'nettoyage',
    'maintenance': 'nettoyage',
    'hôtellerie': 'hotellerie',
    'hotellerie': 'hotellerie',
    'restauration': 'hotellerie',
    'hospitality': 'hotellerie',
    'catering': 'hotellerie',
    'agriculture': 'agriculture',
    'récolte': 'agriculture',
    'recolte': 'agriculture',
    'harvesting': 'agriculture',
    'polyvalent': 'polyvalent',
    'multi': 'polyvalent',
    'accessible rapidement': 'polyvalent',
}
EURES_CANDIDAT_SECTOR_SALARY_FIELDS = {
    'vente': ('tally_q20_salary_type', 'tally_q20_salary_min'),
    'nettoyage': ('tally_q22_salary_type', 'tally_q22_salary_min'),
    'hotellerie': ('tally_q25_salary_type', 'tally_q25_salary_min'),
    'agriculture': ('tally_q27_salary_type', 'tally_q27_salary_min'),
    'polyvalent': ('tally_q29_salary_type', 'tally_q29_salary_min'),
}
EURES_EMPLOYEUR_SECTOR_SALARY_FIELDS = {
    'vente': ('tally_q10_salary_type', 'tally_q10_salary_min', 'tally_q10_salary_max'),
    'nettoyage': ('tally_q11_salary_type', 'tally_q11_salary_min', 'tally_q11_salary_max'),
    'hotellerie': ('tally_q12_salary_type', 'tally_q12_salary_min', 'tally_q12_salary_max'),
    'agriculture': ('tally_q13_salary_type', 'tally_q13_salary_min', 'tally_q13_salary_max'),
    'polyvalent': ('tally_q14_salary_type', 'tally_q14_salary_min', 'tally_q14_salary_max'),
}
EURES_PUBLIC_SECTOR_LABELS = {
    'vente': 'Vente et commerce',
    'nettoyage': 'Nettoyage et entretien',
    'hotellerie': 'Hôtellerie et restauration',
    'agriculture': 'Agriculture et récolte',
    'polyvalent': 'Missions polyvalentes',
}
WIZARD_STATE_KEY = '__wizard_v3_state'
JSON_EXPORT_COLUMNS = {
    'metiers_json',
    'prestations_orp_json',
    'prestations_indirectes_json',
    'prestations_json',
    'catalogue_formations_json',
    'emploi_indicateurs_json',
    'prestations_details_json',
    'finess_json',
}
FIELD_LABELS = {
    'uuid': 'Identifiant de reprise',
    'es_nom': "Nom de l'établissement",
    'es_departement': 'Département',
    'finess_main': 'FINESS principal',
    'validateur_nom': 'Nom du validateur',
    'validateur_prenom': 'Prénom du validateur',
    'validateur_email': 'Email du validateur',
    'saisie_terminee': 'Saisie terminée',
    'check_esrp': 'Dispositif ESRP',
    'check_espo': 'Dispositif ESPO',
    'check_ueros': 'Dispositif UEROS',
    'check_deac': 'Dispositif DEAc',
    'etp_esrp': 'ETP ESRP',
    'etp_espo': 'ETP ESPO',
    'etp_ueros': 'ETP UEROS',
    'etp_deac': 'ETP DEAc',
    'q32_implantation': "Contexte social - Zone d'implantation",
    'q33_transports': 'Contexte social - Transports',
    'q33_pmr': 'Contexte social - Accessibilité PMR',
    'q33_alternatif': 'Contexte social - Transport alternatif',
    'q34_prefecture': 'Contexte social - Préfecture',
    'q35_hebergement': 'Contexte social - Hébergement',
    'q35_places': 'Contexte social - Nombre de places',
    'q35_weekend': 'Contexte social - Hébergement le week-end',
    'q36_restaurant': 'Contexte social - Restauration',
    'q37_cuisine': 'Contexte social - Cuisine',
    'q38_dui': 'Contexte social - DUI',
    'q40_remuneration': 'Contexte social - Rémunération',
    'q40_operateur': 'Contexte social - Opérateur rémunération',
    'autre_dispositif_eval': "Autre dispositif d'évaluation",
    'autre_dispositif_eval_avec_orp_cdaph': "Autre dispositif d'évaluation avec ORP CDAPH",
    'autre_dispositif_eval_sans_orp_cdaph': "Autre dispositif d'évaluation sans ORP CDAPH",
}
COH_LABELS = {
    'genre': ('Genre', ['Hommes', 'Femmes', 'Autre']),
    'age': ('Âge', ['16-17 ans', '18-19 ans', '20-24 ans', '25-29 ans', '30-34 ans', '35-39 ans', '40-44 ans', '45-49 ans', '50-54 ans', '55-59 ans', '60 ans et +', 'Je ne sais pas']),
    'niveau_entree': ("Niveau de formation à l'entrée", ['Niveau 2 (CEP, sans formation)', 'Niveau 3 (CAP, BEP)', 'Niveau 4 (Bac, Bac pro)', 'Niveau 5 (Bac +2, BTS)', 'Niveau 6 (Bac +3, Bac +4)', 'Niveau 7 (Bac +5)', 'Je ne sais pas']),
    'situation_entree': ("Situation à l'entrée", ['En emploi du secteur privé', 'Dont entreprises adaptées', 'Dont alternance (secteur privé)', 'En emploi secteur public', 'FPT', 'Dont contractuels (FPT)', 'Dont alternance (FPT)', 'FPH', 'Dont contractuels (FPH)', 'Dont alternance (FPH)', 'FPE', 'Dont contractuels (FPE)', 'Dont alternance (FPE)', 'En activité non salariée', "Travailleurs d'ESAT", "Sans emploi depuis moins d'1 an", 'Sans emploi depuis 1 à 2 ans', 'Sans emploi depuis 2 ans ou plus', "N'avaient jamais travaillé", 'En formation', 'Je ne sais pas']),
    'ressources_entree': ("Ressources à l'entrée", ['Salaire', 'AAH', 'Allocations chômage', 'RSA', 'Indemnités Journalières', 'Rentes AT/MP / Pension invalidité', 'Autres', 'Aucune', 'Je ne sais pas']),
    'pathologies': ('Pathologies', ['1 pathologie', '2 pathologies', '3 pathologies et plus', 'Non connu']),
    'origine_handicap': ('Origine du handicap', ['Congénitale', 'Maladie', 'Accident de vie privée', 'Accident travail / Pro', 'Maladie professionnelle', 'Autres', 'Non connue']),
    'lesion_origine': ('Origine de la lésion cérébrale (UEROS)', ['Traumatisme crânien', 'AVC', 'Tumeur cérébrale', 'Epilepsie', 'Autres pathologies neuro', 'Non connue']),
}
HANDICAP_TYPES = [
    'Déficiences intellectuelles',
    'Autisme et autres TED',
    'Troubles psychiques',
    'Troubles langage/apprentissages',
    'Déficiences auditives',
    'Déficiences visuelles',
    'Déficiences motrices',
    'Déficiences métaboliques/nutritionnelles',
    'Cérébro-lésions',
    'Polyhandicap',
    'TCC (comportement/com.)',
    'Diagnostics en cours',
    'Autres types de déficiences',
    'Je ne sais pas',
]
EXPORT_KEY_LABELS = {
    'done': 'Statut complété',
    'fileActive': 'File active',
    'preaccueilSansSuite': 'Pré-accueils sans suite',
    'sortiesAvantTerme': 'Sorties définitives avant le terme de la prestation',
    'sortiesTerme': 'Sorties au terme de la prestation',
    'sorties': 'Sorties définitives en 2025',
    'journees': 'Journées réalisées',
    'journeesTheoriques': 'Journées théoriques',
    'enabled': 'Activé',
    'beneficiaires': 'Nombre de personnes bénéficiaires',
    'discontinue_personnes': 'Nombre de personnes accompagnées en discontinue',
    'presentiel_total': 'Présentiel - total',
    'presentiel_complet': 'Présentiel - temps complet',
    'presentiel_partiel': 'Présentiel - temps partiel',
    'hybride_total': 'Hybride - total',
    'hybride_complet': 'Hybride - temps complet',
    'hybride_partiel': 'Hybride - temps partiel',
    'distanciel_total': 'Distanciel - total',
    'distanciel_complet': 'Distanciel - temps complet',
    'distanciel_partiel': 'Distanciel - temps partiel',
    'hors_murs_personnes': 'Hors les murs - personnes accompagnées',
    'hors_murs_journees': 'Hors les murs - journées',
    'hebergees_personnes': 'Hébergement - personnes hébergées',
    'hebergees_journees': 'Hébergement - journées',
    'hebergees_nuitees': 'Hébergement - nuitées',
    'activites_intermediaires_site': 'Activité intermédiaire - sur site',
    'activites_intermediaires_ambulatoire': 'Activité intermédiaire - en ambulatoire',
    'activites_sortie_parcours_site': 'Suite de parcours - sur site',
    'activites_sortie_parcours_ambulatoire': 'Suite de parcours - en ambulatoire',
    'droit_commun': 'Droit commun',
    'sante_social': 'Santé / Social',
    'readaptation_professionnelle': 'Réadaptation professionnelle',
    'readaptation_ueros': 'Réadaptation professionnelle - UEROS',
    'readaptation_espo': 'Réadaptation professionnelle - ESPO',
    'readaptation_esrp': 'Réadaptation professionnelle - ESRP',
    'readaptation_dfa': 'Réadaptation professionnelle - DFA',
    'inconnu': 'Situations inconnues',
    'autre': 'Autre préconisation',
    'autres': 'Autres',
    'autres_precision': 'Précision',
    'je_ne_sais_pas': 'Je ne sais pas',
    'dept': 'Département',
    'count': 'Effectif',
    'principal': 'Principal',
    'associe': 'Associé',
}

# Public mount for the daily capture tool.
try:
    from tools.fagerh_suivi_local.app import app as fagerh_suivi_app
except Exception:
    fagerh_suivi_app = None


def _resolve_form_path(form_id: str, raw_path: str) -> str | None:
    """
    Resolve friendly form URLs.
    Supports:
    - explicit file path (existing)
    - "<name>" -> "<name>.html"
    - "<name>/" -> "<name>/index.html" then "<name>.html"
    """
    safe_form_dir = FORMS_DIR / form_id
    candidate = (raw_path or '').strip().lstrip('/')
    if not candidate:
        return None

    candidates: list[str] = []
    candidates.append(candidate)
    if candidate.endswith('/'):
        base = candidate.rstrip('/')
        candidates.append(f"{base}/index.html")
        candidates.append(f"{base}.html")
    else:
        if '.' not in Path(candidate).name:
            candidates.append(f"{candidate}.html")
            candidates.append(f"{candidate}/index.html")

    for c in candidates:
        if (safe_form_dir / c).is_file():
            return c
    return None


def _require_admin_auth():
    """HTTP Basic auth guard for admin endpoints."""
    username = os.environ.get('ADMIN_USERNAME')
    password = os.environ.get('ADMIN_PASSWORD')
    if not username or not password:
        return Response(
            'Admin access not configured. Set ADMIN_USERNAME and ADMIN_PASSWORD.',
            503,
            {'Content-Type': 'text/plain; charset=utf-8'},
        )

    auth = request.authorization
    if not auth or auth.username != username or auth.password != password:
        return Response(
            'Authentication required',
            401,
            {'WWW-Authenticate': 'Basic realm="Grist Admin"'},
        )
    return None


def admin_required(fn):
    """Decorator to protect admin pages and APIs with basic auth."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        denied = _require_admin_auth()
        if denied:
            return denied
        return fn(*args, **kwargs)
    return wrapper


def get_form_config(form_id: str, role: str | None = None) -> dict:
    """Get configuration for a form from environment variables."""
    form_id_upper = form_id.replace('-', '_').upper()
    role_upper = (role or '').strip().replace('-', '_').upper()

    def _env_with_role(base: str) -> str | None:
        if role_upper:
            value = os.environ.get(f'{base}_{form_id_upper}_{role_upper}')
            if value:
                return value
        return os.environ.get(f'{base}_{form_id_upper}')

    doc_id = _env_with_role('GRIST_DOC')
    if not doc_id:
        return None

    return {
        'doc_id': doc_id,
        'table_id': _env_with_role('GRIST_TABLE') or 'Reponses',
        'api_key': _env_with_role('GRIST_API_KEY') or os.environ.get('GRIST_API_KEY'),
    }


def get_eures_stats_config() -> dict | None:
    """Get configuration for the optional EURES monthly stats table."""
    base = get_form_config('eures-beta', 'candidate') or get_form_config('eures-beta')
    if not base:
        return None

    return {
        'doc_id': base['doc_id'],
        'table_id': os.environ.get('GRIST_TABLE_EURES_BETA_STATS', EURES_STATS_TABLE_DEFAULT),
        'api_key': base.get('api_key'),
    }


def get_eures_matching_config() -> dict | None:
    """Get configuration for EURES beta matching tables."""
    base = get_form_config('eures-beta', 'candidate') or get_form_config('eures-beta')
    if not base:
        return None
    return {
        'doc_id': base['doc_id'],
        'table_id': EURES_MATCHINGS_TABLE,
        'api_key': base.get('api_key'),
    }


def get_eures_invitations_config() -> dict | None:
    """Get configuration for the optional EURES invitations table."""
    base = get_form_config('eures-beta', 'candidate') or get_form_config('eures-beta')
    if not base:
        return None
    return {
        'doc_id': base['doc_id'],
        'table_id': os.environ.get('GRIST_TABLE_EURES_BETA_INVITATIONS', EURES_INVITATIONS_TABLE_DEFAULT),
        'api_key': base.get('api_key'),
    }


def get_brevo_config() -> dict:
    return {
        'api_key': os.environ.get('BREVO_API_KEY', '').strip(),
        'from_email': os.environ.get('BREVO_FROM_EMAIL', '').strip(),
        'from_name': os.environ.get('BREVO_FROM_NAME', 'EURES beta').strip() or 'EURES beta',
    }


def get_brevo_health(check_api: bool = False, timeout: int = 8) -> dict:
    """Return a sanitized Brevo health snapshot suitable for admin and health checks."""
    brevo = get_brevo_config()
    configured = bool(brevo['api_key'] and brevo['from_email'])
    result = {
        'configured': configured,
        'from_email': brevo['from_email'],
        'from_name': brevo['from_name'],
        'api_ok': None,
        'http_status': None,
        'status': 'ok' if configured else 'missing_config',
        'message': 'Brevo configuration complete.' if configured else 'BREVO_API_KEY or BREVO_FROM_EMAIL is missing.',
    }
    if not configured or not check_api:
        return result

    try:
        response = requests.get(
            'https://api.brevo.com/v3/account',
            headers={
                'accept': 'application/json',
                'api-key': brevo['api_key'],
            },
            timeout=timeout,
        )
        result['http_status'] = response.status_code
        if response.ok:
            result['api_ok'] = True
            result['status'] = 'ok'
            result['message'] = 'Brevo API reachable.'
            return result

        result['api_ok'] = False
        result['status'] = 'api_error'
        try:
            payload = response.json()
        except ValueError:
            payload = {}
        result['message'] = str(payload.get('message') or payload.get('error') or f'Brevo API returned HTTP {response.status_code}.')
        return result
    except requests.RequestException as exc:
        result['api_ok'] = False
        result['status'] = 'request_error'
        result['message'] = str(exc)
        return result


def ensure_brevo_ready(check_api: bool = False) -> dict:
    """Raise a readable error when Brevo is not ready for transactional sends."""
    health = get_brevo_health(check_api=check_api)
    if not health['configured']:
        raise RuntimeError('Brevo configuration is incomplete. Missing BREVO_API_KEY or BREVO_FROM_EMAIL.')
    if check_api and health['api_ok'] is not True:
        raise RuntimeError(f"Brevo API check failed: {health['message']}")
    return health


def _now_iso_utc() -> str:
    return datetime.utcnow().isoformat() + 'Z'


def _initial_matching_workflow_status(scoring_status: str) -> str:
    return 'a_valider_admin' if str(scoring_status or '').strip().lower() in {'a_valider', 'auto_envoyable'} else 'calcule'


def _matching_workflow_status(fields: dict) -> str:
    current = str((fields or {}).get('workflow_status') or '').strip().lower()
    if current:
        return current
    if (fields or {}).get('embauche_confirmee_at'):
        return 'embauche_confirmee'
    if (fields or {}).get('mise_en_relation_at'):
        return 'mise_en_relation_faite'
    employer_response = str((fields or {}).get('employer_response') or '').strip().lower()
    if employer_response == 'contact':
        return 'accepte_employeur'
    if employer_response == 'not_contact':
        return 'refuse_employeur'
    if (fields or {}).get('sent_to_employer_at'):
        return 'envoye_employeur'
    admin_status = str((fields or {}).get('admin_status') or '').strip().lower()
    if admin_status == 'accepted':
        return 'envoye_employeur'
    if admin_status == 'refused':
        return 'refuse_admin'
    return _initial_matching_workflow_status((fields or {}).get('statut', ''))


def _matching_workflow_update_fields(target_status: str, actor: str) -> dict:
    now = _now_iso_utc()
    fields = {
        'workflow_status': target_status,
        'workflow_status_updated_at': now,
        'workflow_status_updated_by': actor,
    }
    if target_status == 'envoye_employeur':
        fields['sent_to_employer_at'] = now
        fields['sent_to_employer_by'] = actor
    elif target_status == 'accepte_employeur':
        fields['employer_response'] = 'contact'
        fields['employer_response_at'] = now
        fields['employer_response_source'] = 'email_cta'
    elif target_status == 'refuse_employeur':
        fields['employer_response'] = 'not_contact'
        fields['employer_response_at'] = now
        fields['employer_response_source'] = 'email_cta'
    elif target_status == 'mise_en_relation_faite':
        fields['mise_en_relation_at'] = now
        fields['mise_en_relation_by'] = actor
    elif target_status == 'embauche_confirmee':
        fields['embauche_confirmee_at'] = now
        fields['embauche_confirmee_by'] = actor
    return fields


def _matching_workflow_label(status: str) -> str:
    return {
        'calcule': 'calculé',
        'a_valider_admin': 'à valider',
        'refuse_admin': 'refus admin',
        'valide_admin': 'validé admin',
        'envoye_employeur': 'envoyé employeur',
        'accepte_employeur': 'accepté employeur',
        'refuse_employeur': 'refusé employeur',
        'mise_en_relation_faite': 'mise en relation faite',
        'embauche_confirmee': 'embauche confirmée',
    }.get(str(status or '').strip().lower(), str(status or '').strip() or 'inconnu')


def get_eures_mail_signature_name() -> str:
    return os.environ.get('EURES_SIGNATURE_NAME', 'Eric Barthélémy').strip() or 'Eric Barthélémy'


def get_public_app_base_url() -> str:
    configured = os.environ.get('PUBLIC_APP_BASE_URL', '').strip()
    if configured:
        return configured.rstrip('/')
    if request and request.url_root:
        return request.url_root.rstrip('/')
    return 'https://eures-beta.osc-fr1.scalingo.io'


def get_eures_email_action_serializer() -> URLSafeSerializer:
    secret = (
        os.environ.get('EURES_EMAIL_ACTION_SECRET', '').strip()
        or os.environ.get('ADMIN_PASSWORD', '').strip()
        or os.environ.get('BREVO_API_KEY', '').strip()
    )
    if not secret:
        raise RuntimeError('Missing EURES email action secret configuration.')
    return URLSafeSerializer(secret_key=secret, salt='eures-email-action')


def _generate_eures_invitation_token() -> str:
    return secrets.token_urlsafe(24)


def _normalize_eures_invitation_role(value: str) -> str:
    role = str(value or '').strip().lower().replace('-', '_')
    if role in {'candidat', 'candidate'}:
        return 'candidate'
    if role in {'employeur', 'employer'}:
        return 'employer'
    return ''


def _normalize_eures_invitation_status(value: str) -> str:
    status = str(value or '').strip().lower().replace(' ', '_')
    if status in EURES_INVITATION_ALLOWED_STATUSES:
        return status
    return ''


def _build_eures_invitation_link(role: str, language: str, invite_token: str) -> str:
    page = 'candidate' if role == 'candidate' else 'employer'
    lang = str(language or 'fr').strip().lower()
    if lang not in {'fr', 'en', 'de'}:
        lang = 'fr'
    base = get_public_app_base_url()
    return f'{base}/forms/eures-beta/{page}?lang={lang}&invite_token={invite_token}'


def _eures_invitation_lookup_key(role: str, email: str) -> tuple[str, str]:
    return (_normalize_eures_invitation_role(role), normalize_email(email))


def _extract_csv_rows(csv_text: str) -> list[dict]:
    raw_text = str(csv_text or '').replace('\r\n', '\n').replace('\r', '\n').lstrip('\ufeff')
    if not raw_text.strip():
        return []

    sample = '\n'.join(raw_text.splitlines()[:5])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;')
    except Exception:
        dialect = csv.excel

    buffer = StringIO(raw_text)
    reader = csv.DictReader(buffer, dialect=dialect)
    if reader.fieldnames:
        reader.fieldnames = [str(name or '').strip().lstrip('\ufeff') for name in reader.fieldnames]

    rows = []
    for row in reader:
        if not isinstance(row, dict):
            continue
        cleaned = {
            str(key or '').strip().lstrip('\ufeff'): str(value or '').strip()
            for key, value in row.items()
            if str(key or '').strip().lstrip('\ufeff')
        }
        if any(cleaned.values()):
            rows.append(cleaned)
    return rows


def _coalesce_row_value(row: dict, *keys: str) -> str:
    for key in keys:
        if key in row and str(row.get(key) or '').strip():
            return str(row.get(key) or '').strip()
    return ''


def _parse_eures_invitation_rows(payload: dict) -> list[dict]:
    rows = payload.get('rows')
    if isinstance(rows, list):
        return [row for row in rows if isinstance(row, dict)]

    csv_text = payload.get('csv_text')
    if isinstance(csv_text, str) and csv_text.strip():
        return _extract_csv_rows(csv_text)

    return []


def _normalize_eures_invitation_row(row: dict, actor: str, batch_id: str) -> dict:
    role = _normalize_eures_invitation_role(_coalesce_row_value(row, 'role', 'type', 'profil', 'profile'))
    email = normalize_email(_coalesce_row_value(row, 'email', 'mail', 'contact_email'))
    language = _coalesce_row_value(row, 'language', 'langue', 'lang') or 'fr'
    invitation_status = _normalize_eures_invitation_status(
        _coalesce_row_value(row, 'invitation_status', 'status', 'statut')
    ) or 'invitation_a_envoyer'
    token = _coalesce_row_value(row, 'invite_token', 'token') or _generate_eures_invitation_token()
    return {
        'role': role,
        'email': email,
        'first_name': _coalesce_row_value(row, 'first_name', 'prenom', 'prénom'),
        'last_name': _coalesce_row_value(row, 'last_name', 'nom'),
        'company_name': _coalesce_row_value(row, 'company_name', 'entreprise', 'company'),
        'language': language.lower(),
        'source': _coalesce_row_value(row, 'source', 'origine') or 'csv_import',
        'external_ref': _coalesce_row_value(row, 'external_ref', 'external_id', 'id_externe', 'id'),
        'invite_token': token,
        'invite_link': _build_eures_invitation_link(role, language, token) if role else '',
        'invitation_status': invitation_status,
        'invitation_status_updated_at': _now_iso_utc(),
        'invitation_status_updated_by': actor,
        'import_batch_id': batch_id,
        'imported_at': _now_iso_utc(),
        'imported_by': actor,
        'notes': _coalesce_row_value(row, 'notes', 'note', 'comment', 'commentaire'),
    }


def list_eures_invitations() -> list[dict]:
    config = get_eures_invitations_config()
    if not config:
        raise RuntimeError('EURES invitations configuration is incomplete.')
    headers = _eures_admin_headers(config)
    records = fetch_table_records(config['doc_id'], config['table_id'], headers)
    rows = []
    for rec in records:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        rows.append({
            'record_id': rec.get('id'),
            'role': fields.get('role', ''),
            'email': fields.get('email', ''),
            'first_name': fields.get('first_name', ''),
            'last_name': fields.get('last_name', ''),
            'company_name': fields.get('company_name', ''),
            'language': fields.get('language', ''),
            'source': fields.get('source', ''),
            'external_ref': fields.get('external_ref', ''),
            'invite_token': fields.get('invite_token', ''),
            'invite_link': fields.get('invite_link', ''),
            'invitation_status': fields.get('invitation_status', 'invitation_a_envoyer'),
            'sent_at': fields.get('sent_at', ''),
            'answered_at': fields.get('answered_at', ''),
            'linked_form_role': fields.get('linked_form_role', ''),
            'linked_record_id': fields.get('linked_record_id', ''),
            'linked_record_key': fields.get('linked_record_key', ''),
            'matching_status': fields.get('matching_status', ''),
            'notes': fields.get('notes', ''),
            'import_batch_id': fields.get('import_batch_id', ''),
            'imported_at': fields.get('imported_at', ''),
            'imported_by': fields.get('imported_by', ''),
        })
    rows.sort(key=lambda row: (str(row.get('invitation_status') or ''), -int(row.get('record_id') or 0)))
    return rows


def upsert_eures_invitation_rows(rows: list[dict], actor: str) -> dict:
    config = get_eures_invitations_config()
    if not config:
        raise RuntimeError('EURES invitations configuration is incomplete.')
    headers = _eures_admin_headers(config)
    ensure_table_columns(config, EURES_INVITATION_FIELDS, headers)
    existing_records = fetch_table_records(config['doc_id'], config['table_id'], headers)
    by_key: dict[tuple[str, str], dict] = {}
    for rec in existing_records:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        key = _eures_invitation_lookup_key(fields.get('role', ''), fields.get('email', ''))
        if key[0] and key[1]:
            by_key[key] = rec

    batch_id = f'invite-import-{int(time.time())}'
    to_create = []
    to_update = []
    skipped = []
    for raw_row in rows:
        normalized = _normalize_eures_invitation_row(raw_row, actor=actor, batch_id=batch_id)
        role = normalized['role']
        email = normalized['email']
        if role not in EURES_INVITATION_ALLOWED_ROLES or not email:
            skipped.append({
                'row': raw_row,
                'reason': 'missing_or_invalid_role_or_email',
            })
            continue
        existing = by_key.get((role, email))
        if existing:
            existing_fields = existing.get('fields', {}) if isinstance(existing.get('fields'), dict) else {}
            merged = dict(existing_fields)
            merged.update({
                key: value for key, value in normalized.items()
                if value not in (None, '')
            })
            merged['invite_token'] = existing_fields.get('invite_token') or normalized['invite_token']
            merged['invite_link'] = _build_eures_invitation_link(role, merged.get('language', 'fr'), merged['invite_token'])
            to_update.append({'id': existing['id'], 'fields': merged})
        else:
            to_create.append({'fields': normalized})

    base_url = f"{GRIST_BASE_URL}/api/docs/{config['doc_id']}/tables/{config['table_id']}/records"
    if to_create:
        resp = write_grist_records('POST', base_url, {'records': to_create}, headers)
        if resp.status_code != 200:
            raise RuntimeError(f'Failed to create invitations: HTTP {resp.status_code} - {resp.text}')
    if to_update:
        resp = write_grist_records('PATCH', base_url, {'records': to_update}, headers)
        if resp.status_code != 200:
            raise RuntimeError(f'Failed to update invitations: HTTP {resp.status_code} - {resp.text}')

    return {
        'ok': True,
        'batch_id': batch_id,
        'created': len(to_create),
        'updated': len(to_update),
        'skipped': skipped,
    }


def update_eures_invitation_record_by_id(record_id: int, fields: dict, headers: dict | None = None):
    """Patch one invitation record by Grist record id."""
    config = get_eures_invitations_config()
    if not config:
        raise RuntimeError('EURES invitations configuration is incomplete.')
    if headers is None:
        headers = _eures_admin_headers(config)
    ensure_table_columns(config, set(fields.keys()) & EURES_INVITATION_FIELDS, headers)
    allowed_columns = get_table_columns(config, headers)
    filtered_fields = {k: v for k, v in fields.items() if k in allowed_columns}
    base_url = f"{GRIST_BASE_URL}/api/docs/{config['doc_id']}/tables/{config['table_id']}/records"
    resp = write_grist_records('PATCH', base_url, {'records': [{'id': record_id, 'fields': filtered_fields}]}, headers)
    if resp.status_code != 200:
        raise RuntimeError(f'Failed to update invitation: HTTP {resp.status_code} - {resp.text}')
    return resp


def find_eures_invitation_by_token(invite_token: str, headers: dict | None = None) -> dict | None:
    """Find one invitation row by invite token."""
    token = str(invite_token or '').strip()
    if not token:
        return None
    config = get_eures_invitations_config()
    if not config:
        raise RuntimeError('EURES invitations configuration is incomplete.')
    if headers is None:
        headers = _eures_admin_headers(config)
    records = fetch_table_records(config['doc_id'], config['table_id'], headers)
    for rec in records:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        if str(fields.get('invite_token') or '').strip() == token:
            return rec
    return None


def find_eures_invitation_by_role_email(role: str, email: str, headers: dict | None = None) -> dict | None:
    """Find one invitation row by normalized role and email."""
    normalized_role = _normalize_eures_invitation_role(role)
    normalized_email = normalize_email(email)
    if not normalized_role or not normalized_email:
        return None
    config = get_eures_invitations_config()
    if not config:
        raise RuntimeError('EURES invitations configuration is incomplete.')
    if headers is None:
        headers = _eures_admin_headers(config)
    records = fetch_table_records(config['doc_id'], config['table_id'], headers)
    for rec in records:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        if _eures_invitation_lookup_key(fields.get('role', ''), fields.get('email', '')) == (normalized_role, normalized_email):
            return rec
    return None


def link_eures_invitation_after_save(role: str, request_fields: dict, saved_record: dict, matching_result: dict | None = None) -> dict:
    """Link a saved EURES questionnaire to its invitation when possible."""
    if _normalize_eures_invitation_role(role) not in EURES_INVITATION_ALLOWED_ROLES:
        return {'linked': False, 'reason': 'unsupported_role'}

    config = get_eures_invitations_config()
    if not config:
        return {'linked': False, 'reason': 'invitations_not_configured'}

    headers = _eures_admin_headers(config)
    invite_token = str((request_fields or {}).get('invite_token') or '').strip()
    saved_fields = saved_record.get('fields', {}) if isinstance(saved_record.get('fields'), dict) else {}
    email = normalize_email(
        (request_fields or {}).get('email')
        or saved_fields.get('email')
        or saved_fields.get('tally_q18')
        or saved_fields.get('tally_q33')
    )

    invitation = None
    link_mode = ''
    if invite_token:
        invitation = find_eures_invitation_by_token(invite_token, headers=headers)
        if invitation:
            link_mode = 'token'

    if invitation is None and email:
        invitation = find_eures_invitation_by_role_email(role, email, headers=headers)
        if invitation:
            link_mode = 'email'

    if not invitation:
        return {
            'linked': False,
            'reason': 'invitation_not_found',
            'invite_token_present': bool(invite_token),
            'email_present': bool(email),
        }

    record_id = invitation.get('id')
    if not record_id:
        return {'linked': False, 'reason': 'invalid_invitation_record'}

    saved_record_key = str(saved_fields.get('id_tally') or saved_fields.get('uuid') or '').strip()
    now = _now_iso_utc()
    invitation_fields = invitation.get('fields', {}) if isinstance(invitation.get('fields'), dict) else {}
    notes = str(invitation_fields.get('notes') or '').strip()
    if link_mode == 'email':
        fallback_note = 'Lien questionnaire/invitation rapproche via email.'
        if fallback_note not in notes:
            notes = f'{notes} | {fallback_note}'.strip(' |')

    update_fields = {
        'answered_at': now,
        'linked_form_role': _normalize_eures_invitation_role(role),
        'linked_record_id': str(saved_record.get('id') or ''),
        'linked_record_key': saved_record_key,
        'invitation_status': 'questionnaire_recu',
        'invitation_status_updated_at': now,
        'invitation_status_updated_by': 'system_questionnaire_submission',
        'notes': notes,
    }
    if isinstance(matching_result, dict) and matching_result.get('processed'):
        update_fields['matching_status'] = 'matching_calcule'
        update_fields['matching_status_updated_at'] = now
        update_fields['matching_status_updated_by'] = 'system_matching'

    update_eures_invitation_record_by_id(int(record_id), update_fields, headers=headers)
    return {
        'linked': True,
        'mode': link_mode,
        'invitation_record_id': record_id,
        'linked_record_id': saved_record.get('id'),
        'linked_record_key': saved_record_key,
        'email': email,
    }


def build_brevo_invitation_email(invitation_row: dict) -> tuple[str, str, str, str]:
    """Build recipient, subject, text body and HTML body for one invitation."""
    role = _normalize_eures_invitation_role(invitation_row.get('role', ''))
    if role not in EURES_INVITATION_ALLOWED_ROLES:
        raise RuntimeError('Invitation role is missing or invalid.')
    recipient = normalize_email(invitation_row.get('email', ''))
    if not recipient:
        raise RuntimeError('Invitation email is missing.')

    language = str(invitation_row.get('language') or 'fr').strip().lower()
    if language not in {'fr', 'en', 'de'}:
        language = 'fr'
    first_name = str(invitation_row.get('first_name') or '').strip()
    company_name = str(invitation_row.get('company_name') or '').strip()
    invite_token = str(invitation_row.get('invite_token') or '').strip() or _generate_eures_invitation_token()
    invite_link = str(invitation_row.get('invite_link') or '').strip() or _build_eures_invitation_link(role, language, invite_token)
    signature_name = get_eures_mail_signature_name()

    if role == 'employer':
        if language == 'en':
            subject = "[EURES beta] Recruit more easily through European mobility"
            preheader = "Experimental service: describe your hiring need in less than 5 minutes."
            title = "Are you having trouble recruiting?"
            body_lines = [
                "EURES beta is an initiative led by EURES and France Travail to support recruitment and professional mobility in the Greater Region.",
                "In just a few minutes, describe your hiring need and access a wider pool of candidates open to European mobility.",
                "The more precise your need is, the easier it is for us to identify relevant profiles.",
            ]
            cta = "Complete the questionnaire"
            cta_note = "Estimated time: less than 5 minutes."
            footer = "EURES beta is an initiative led by EURES and France Travail to support recruitment and professional mobility in the Greater Region."
        elif language == 'de':
            subject = "[EURES beta] Einfacher rekrutieren dank europäischer Mobilität"
            preheader = "Experimenteller Service: Beschreiben Sie Ihren Bedarf in weniger als 5 Minuten."
            title = "Haben Sie Schwierigkeiten bei der Rekrutierung?"
            body_lines = [
                "EURES beta ist eine von EURES und France Travail getragene Initiative zur Unterstützung von Rekrutierung und beruflicher Mobilität in der Großregion.",
                "Beschreiben Sie in wenigen Minuten Ihren Bedarf und erhalten Sie Zugang zu einem größeren Pool von Kandidatinnen und Kandidaten, die für europäische Mobilität offen sind.",
                "Je präziser Ihr Bedarf ist, desto besser können passende Profile identifiziert werden.",
            ]
            cta = "Fragebogen ausfüllen"
            cta_note = "Geschätzte Dauer: weniger als 5 Minuten."
            footer = "EURES beta ist eine von EURES und France Travail getragene Initiative zur Unterstützung von Rekrutierung und beruflicher Mobilität in der Großregion."
        else:
            subject = "[EURES beta] Vous avez des difficultés à recruter ?"
            preheader = "Expérimentation en cours : décrivez votre besoin en moins de 5 minutes."
            title = "Vous avez des difficultés à recruter ?"
            body_lines = [
                "EURES beta est une expérimentation portée par EURES et France Travail pour faciliter les recrutements et la mobilité professionnelle dans la Grande Région.",
                "En quelques minutes, décrivez votre besoin et accédez à un vivier plus large de candidats ouverts à la mobilité européenne.",
                "Plus votre besoin est précis, plus nous serons en mesure d'identifier des profils susceptibles de correspondre à votre recherche.",
            ]
            cta = "Compléter le questionnaire"
            cta_note = "Temps estimé : moins de 5 minutes."
            footer = "EURES beta est une expérimentation portée par EURES et France Travail pour faciliter les recrutements et la mobilité professionnelle dans la Grande Région."

        text_body = (
            f"{title}\n\n"
            f"{preheader}\n\n"
            + "\n\n".join(body_lines)
            + f"\n\n{cta}: {invite_link}\n\n{cta_note}\n\n{footer}\n"
        )
        body_html = f"""
<!doctype html>
<html lang="{escape(language)}">
  <body style="margin:0;padding:0;background:#f6f7fb;font-family:Arial,'Helvetica Neue',sans-serif;color:#16253d;">
    <div style="display:none;max-height:0;overflow:hidden;opacity:0;mso-hide:all;">
      {escape(preheader)}
    </div>
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f6f7fb;padding:20px 12px;">
      <tr>
        <td align="center">
          <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:560px;background:#ffffff;border:1px solid #dfe6f2;border-radius:20px;overflow:hidden;">
            <tr>
              <td style="padding:24px 24px 8px;">
                <div style="font-size:12px;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;color:#4d678c;">EURES beta</div>
              </td>
            </tr>
            <tr>
              <td style="padding:0 24px 8px;">
                <h1 style="margin:0;font-size:30px;line-height:1.12;font-weight:700;color:#16253d;">{escape(title)}</h1>
              </td>
            </tr>
            <tr>
              <td style="padding:0 24px 4px;">
                <p style="margin:0 0 14px;font-size:16px;line-height:1.55;color:#324765;">{escape(body_lines[0])}</p>
                <p style="margin:0 0 14px;font-size:16px;line-height:1.55;color:#324765;">{escape(body_lines[1])}</p>
                <p style="margin:0 0 22px;font-size:16px;line-height:1.55;color:#324765;">{escape(body_lines[2])}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:0 24px 12px;">
                <a href="{escape(invite_link)}" style="display:block;width:100%;box-sizing:border-box;padding:15px 18px;border-radius:14px;background:#0a66c2;color:#ffffff;text-decoration:none;font-size:16px;font-weight:700;text-align:center;">{escape(cta)}</a>
              </td>
            </tr>
            <tr>
              <td style="padding:0 24px 24px;">
                <p style="margin:0;font-size:14px;line-height:1.5;color:#627892;">{escape(cta_note)}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:18px 24px 24px;border-top:1px solid #e7ecf4;">
                <p style="margin:0;font-size:13px;line-height:1.55;color:#627892;">{escape(footer)}</p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
""".strip()
        return recipient, subject, text_body, body_html, invite_token, invite_link

    if language == 'en':
        subject = "[EURES beta] What if your next opportunity was elsewhere in Europe?"
        preheader = "Experimental service: answer a few questions and explore opportunities through European mobility."
        title = "Explore new professional opportunities in Europe"
        body_lines = [
            "EURES beta is an initiative led by EURES and France Travail to support professional mobility in the Greater Region.",
            "Are you open to a professional opportunity in Luxembourg, in a neighbouring country or elsewhere in Europe? EURES beta helps you discover opportunities aligned with your profile and your mobility plans.",
            "In just a few minutes, describe your experience, skills and preferences so we can identify opportunities that may interest you.",
        ]
        cta = "Discover opportunities"
        cta_note = "Estimated time: less than 5 minutes."
        footer = "EURES beta is an initiative led by EURES and France Travail to support professional mobility in the Greater Region."
    elif language == 'de':
        subject = "[EURES beta] Was wäre, wenn Ihre nächste Chance anderswo in Europa läge?"
        preheader = "Experimenteller Service: Beantworten Sie ein paar Fragen und entdecken Sie Möglichkeiten durch europäische Mobilität."
        title = "Entdecken Sie neue berufliche Chancen in Europa"
        body_lines = [
            "EURES beta ist eine von EURES und France Travail getragene Initiative zur Unterstützung der beruflichen Mobilität in der Großregion.",
            "Sind Sie offen für eine berufliche Erfahrung in Luxemburg, in einem Nachbarland oder anderswo in Europa? EURES beta hilft Ihnen, Chancen zu entdecken, die zu Ihrem Profil und Ihrem Mobilitätsprojekt passen.",
            "Beschreiben Sie in wenigen Minuten Ihre Erfahrung, Ihre Kompetenzen und Ihre Präferenzen, damit passende Möglichkeiten identifiziert werden können.",
        ]
        cta = "Chancen entdecken"
        cta_note = "Geschätzte Dauer: weniger als 5 Minuten."
        footer = "EURES beta ist eine von EURES und France Travail getragene Initiative zur Unterstützung der beruflichen Mobilität in der Großregion."
    else:
        subject = "[EURES beta] Et si votre prochaine opportunité se trouvait ailleurs en Europe ?"
        preheader = "Expérimentation en cours : répondez à quelques questions et explorez les possibilités offertes par la mobilité européenne."
        title = "Explorez de nouvelles opportunités professionnelles en Europe"
        body_lines = [
            "EURES beta est une expérimentation portée par EURES et France Travail pour faciliter la mobilité professionnelle dans la Grande Région.",
            "Vous êtes ouvert à une expérience professionnelle au Luxembourg, dans un pays voisin ou ailleurs en Europe ? EURES beta vous aide à découvrir des opportunités adaptées à votre profil et à votre projet de mobilité.",
            "En quelques minutes, décrivez votre expérience, vos compétences et vos préférences pour nous permettre d’identifier des opportunités susceptibles de vous intéresser.",
        ]
        cta = "Découvrir les opportunités"
        cta_note = "Temps estimé : moins de 5 minutes."
        footer = "EURES beta est une expérimentation portée par EURES et France Travail pour faciliter la mobilité professionnelle dans la Grande Région."

    text_body = (
        f"{title}\n\n"
        f"{preheader}\n\n"
        + "\n\n".join(body_lines)
        + f"\n\n{cta}: {invite_link}\n\n{cta_note}\n\n{footer}\n"
    )
    html_body = f"""
<!doctype html>
<html lang="{escape(language)}">
  <body style="margin:0;padding:0;background:#f6f7fb;font-family:Arial,'Helvetica Neue',sans-serif;color:#16253d;">
    <div style="display:none;max-height:0;overflow:hidden;opacity:0;mso-hide:all;">
      {escape(preheader)}
    </div>
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f6f7fb;padding:20px 12px;">
      <tr>
        <td align="center">
          <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:560px;background:#ffffff;border:1px solid #dfe6f2;border-radius:20px;overflow:hidden;">
            <tr>
              <td style="padding:24px 24px 8px;">
                <div style="font-size:12px;font-weight:700;letter-spacing:0.08em;text-transform:uppercase;color:#4d678c;">EURES beta</div>
              </td>
            </tr>
            <tr>
              <td style="padding:0 24px 8px;">
                <h1 style="margin:0;font-size:30px;line-height:1.12;font-weight:700;color:#16253d;">{escape(title)}</h1>
              </td>
            </tr>
            <tr>
              <td style="padding:0 24px 4px;">
                <p style="margin:0 0 14px;font-size:16px;line-height:1.55;color:#324765;">{escape(body_lines[0])}</p>
                <p style="margin:0 0 14px;font-size:16px;line-height:1.55;color:#324765;">{escape(body_lines[1])}</p>
                <p style="margin:0 0 22px;font-size:16px;line-height:1.55;color:#324765;">{escape(body_lines[2])}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:0 24px 12px;">
                <a href="{escape(invite_link)}" style="display:block;width:100%;box-sizing:border-box;padding:15px 18px;border-radius:14px;background:#0a66c2;color:#ffffff;text-decoration:none;font-size:16px;font-weight:700;text-align:center;">{escape(cta)}</a>
              </td>
            </tr>
            <tr>
              <td style="padding:0 24px 24px;">
                <p style="margin:0;font-size:14px;line-height:1.5;color:#627892;">{escape(cta_note)}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:18px 24px 24px;border-top:1px solid #e7ecf4;">
                <p style="margin:0;font-size:13px;line-height:1.55;color:#627892;">{escape(footer)}</p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
""".strip()
    return recipient, subject, text_body, html_body, invite_token, invite_link


def get_table_columns(config: dict, headers: dict) -> set[str]:
    """Fetch and cache table column ids to avoid sending unknown fields to Grist."""
    cache_key = (str(config.get('doc_id') or ''), str(config.get('table_id') or ''))
    cached = _TABLE_COLUMNS_CACHE.get(cache_key)
    now = time.time()
    if cached and (now - float(cached.get('loaded_at', 0))) < _TABLE_COLUMNS_CACHE_TTL_SECONDS:
        return set(cached.get('columns', set()))

    url = f"{GRIST_BASE_URL}/api/docs/{config['doc_id']}/tables/{config['table_id']}/columns"
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200:
        raise RuntimeError(f'Failed to read table columns: HTTP {resp.status_code} - {resp.text}')
    payload = resp.json()
    columns = set()
    for col in payload.get('columns', []):
        col_id = (col or {}).get('id')
        if col_id:
            columns.add(str(col_id))
    _TABLE_COLUMNS_CACHE[cache_key] = {
        'columns': columns,
        'loaded_at': now,
    }
    return columns


def ensure_table_columns(config: dict, columns: set[str], headers: dict):
    """Create missing text columns in a Grist table, then refresh cache."""
    if not columns:
        return
    existing = get_table_columns(config, headers)
    missing = sorted(col for col in columns if col not in existing)
    if not missing:
        return

    url = f"{GRIST_BASE_URL}/api/docs/{config['doc_id']}/tables/{config['table_id']}/columns"
    payload = {
        'columns': [
            {
                'id': column_id,
                'fields': {'label': column_id},
                'type': 'Text',
            }
            for column_id in missing
        ]
    }
    resp = requests.post(url, json=payload, headers=headers)
    if resp.status_code != 200:
        raise RuntimeError(f'Failed to create columns in {config["table_id"]}: HTTP {resp.status_code} - {resp.text}')

    cache_key = (str(config.get('doc_id') or ''), str(config.get('table_id') or ''))
    _TABLE_COLUMNS_CACHE.pop(cache_key, None)


def _as_bool(value) -> bool:
    """Normalize truthy values from Grist fields."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    txt = str(value).strip().lower()
    return txt in {'1', 'true', 'vrai', 'oui', 'yes'}


def _has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip() != ''


def _safe_json(value, fallback):
    if isinstance(value, str):
        txt = value.strip()
        if not txt:
            return fallback
        try:
            return json.loads(txt)
        except Exception:
            return fallback
    return value if value is not None else fallback


def _details_payload(fields: dict) -> dict:
    """Return parsed prestations_details_json payload as dict."""
    payload = _safe_json((fields or {}).get('prestations_details_json'), {})
    return payload if isinstance(payload, dict) else {}


def _details_get(payload: dict, path: tuple[str, ...], fallback=None):
    """Safely navigate a nested dict payload."""
    cur = payload
    for key in path:
        if not isinstance(cur, dict):
            return fallback
        cur = cur.get(key)
    return cur if cur is not None else fallback


def _has_any_checked(data: dict) -> bool:
    if not isinstance(data, dict):
        return False
    for v in data.values():
        if isinstance(v, dict):
            if _has_any_checked(v):
                return True
        elif _as_bool(v):
            return True
    return False


def _has_any_number_data(data) -> bool:
    if isinstance(data, dict):
        return any(_has_any_number_data(v) for v in data.values())
    if isinstance(data, list):
        return any(_has_any_number_data(v) for v in data)
    if data is None:
        return False
    txt = str(data).strip()
    if txt == '':
        return False
    try:
        return float(txt) >= 0
    except Exception:
        return False


def _format_conditional_display_name(raw_name: str) -> str:
    """Short admin-friendly label for conditional steps."""
    raw = str(raw_name or '').strip()
    if not raw:
        return ''
    parts = [part.strip() for part in raw.split(' - ') if part.strip()]
    if not parts:
        return raw
    if parts[0] == 'Directes ORP CDAPH':
        return 'Prestations directes: ' + ' > '.join(parts[1:]) if len(parts) > 1 else 'Prestations directes'
    if parts[0] == 'Directes hors ORP CDAPH':
        return 'Prestations directes sans ORP CDAPH: ' + ' > '.join(parts[1:]) if len(parts) > 1 else 'Prestations directes sans ORP CDAPH'
    if parts[0] == 'Indirectes':
        return 'Prestations indirectes: ' + ' > '.join(parts[1:]) if len(parts) > 1 else 'Prestations indirectes'
    return raw


def _split_export_path(path: str) -> list[str]:
    """Split flattened paths while keeping list indexes with their segment."""
    return [part for part in str(path or '').split('.') if part]


def _array_index(segment: str) -> tuple[str, int | None]:
    if '[' not in segment or not segment.endswith(']'):
        return segment, None
    name, _, raw_index = segment.partition('[')
    try:
        return name, int(raw_index[:-1]) - 1
    except Exception:
        return name, None


def humanize_export_field(key: str) -> str:
    """Label top-level Grist/export fields with user-facing wording where known."""
    raw = str(key or '')
    if raw in FIELD_LABELS:
        return FIELD_LABELS[raw]
    if raw.startswith('q') and '_' in raw:
        return raw.replace('_', ' ').upper()
    return raw.replace('_', ' ')


def humanize_export_path(path: str) -> str:
    """Turn flattened JSON paths into labels close to the form wording."""
    raw = str(path or '').strip()
    if not raw:
        return ''
    parts = _split_export_path(raw)
    if not parts:
        return humanize_export_field(raw)

    if parts[0] == 'coh' and len(parts) >= 2:
        block_key, index = _array_index(parts[1])
        title, labels = COH_LABELS.get(block_key, (humanize_export_field(block_key), []))
        if index is not None and 0 <= index < len(labels):
            return f'{title} - {labels[index]}'
        return title

    first_name, first_index = _array_index(parts[0])

    if first_name == 'geoRows' and len(parts) >= 2:
        _, index = _array_index(parts[0])
        suffix = EXPORT_KEY_LABELS.get(parts[1], humanize_export_field(parts[1]))
        prefix = f'Géographie ligne {index + 1}' if index is not None else 'Géographie'
        return f'{prefix} - {suffix}'

    if first_name == 'handicapMatrix' and len(parts) >= 2:
        handicap = HANDICAP_TYPES[first_index] if first_index is not None and 0 <= first_index < len(HANDICAP_TYPES) else 'Type de handicap'
        col = EXPORT_KEY_LABELS.get(parts[1], humanize_export_field(parts[1]))
        return f'Type de handicap principal et associé - {handicap} - {col}'

    if parts[0] in {'directAvecOrp', 'directSansOrp', 'indirect', 'pecFileActive', 'preconisationsBloc', 'sortiesBloc', 'suspensionsBloc', 'orienteursBloc', 'formationsSelection'}:
        group_labels = {
            'directAvecOrp': 'Prestations directes avec ORP CDAPH',
            'directSansOrp': 'Prestations directes sans ORP CDAPH',
            'indirect': 'Prestations indirectes',
            'pecFileActive': 'File active PEC',
            'preconisationsBloc': 'Préconisations lors de la sortie définitive',
            'sortiesBloc': 'Sortie définitive avant le terme de la prestation',
            'suspensionsBloc': 'Suspensions',
            'orienteursBloc': 'Orienteurs',
            'formationsSelection': 'Sélection des formations',
        }
        cleaned = []
        for part in parts[1:]:
            name, index = _array_index(part)
            if name in {'row', 'rows', 'raisons'}:
                continue
            label = EXPORT_KEY_LABELS.get(name, humanize_export_field(name))
            if index is not None:
                label = f'{label} {index + 1}'
            cleaned.append(label)
        return ' - '.join([group_labels[parts[0]], *cleaned])

    if len(parts) == 1:
        name, index = _array_index(parts[0])
        label = EXPORT_KEY_LABELS.get(name, humanize_export_field(name))
        return f'{label} {index + 1}' if index is not None else label

    return ' - '.join(EXPORT_KEY_LABELS.get(_array_index(part)[0], humanize_export_field(_array_index(part)[0])) for part in parts)


def should_skip_readable_export_path(path: str) -> bool:
    """Hide internal navigation/state fields that are not useful in a user-readable export."""
    first = _array_index(_split_export_path(path)[0])[0] if _split_export_path(path) else ''
    return first in {
        'visitedBlocks',
        'currentConditionalId',
        'focusedConditionalStepId',
        'focusedConditionalBlockKey',
        'autoCollapsePrimaryAfterPrestations',
    }


def _xlsx_safe_value(value):
    """Return a spreadsheet-friendly scalar without raw JSON objects."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Oui' if value else 'Non'
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _flatten_for_xlsx(value, prefix=''):
    """Flatten nested JSON as readable path/value rows."""
    rows = []
    if isinstance(value, dict):
        if not value:
            rows.append((prefix, ''))
        for key, child in value.items():
            path = f'{prefix}.{key}' if prefix else str(key)
            rows.extend(_flatten_for_xlsx(child, path))
    elif isinstance(value, list):
        if not value:
            rows.append((prefix, ''))
        for idx, child in enumerate(value, start=1):
            path = f'{prefix}[{idx}]' if prefix else f'[{idx}]'
            rows.extend(_flatten_for_xlsx(child, path))
    else:
        rows.append((prefix, _xlsx_safe_value(value)))
    return rows


def _append_table(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([_xlsx_safe_value(cell) for cell in row])
    for cell in ws[1]:
        cell.style = 'Headline 4'
    ws.freeze_panes = 'A2'
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in column_cells)
        width = min(max(max_len + 2, 12), 80)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def _new_sheet(workbook, title):
    safe = ''.join(ch for ch in str(title or 'Feuille') if ch not in r'[]:*?/\\')[:31] or 'Feuille'
    if safe in workbook.sheetnames:
        base = safe[:27]
        i = 2
        while f'{base} {i}' in workbook.sheetnames:
            i += 1
        safe = f'{base} {i}'
    return workbook.create_sheet(safe)


def build_readable_xlsx(payload: dict) -> BytesIO:
    """Build a human-readable workbook from the current form payload, leaving Grist JSON intact."""
    from openpyxl import Workbook

    fields = payload.get('fields') if isinstance(payload, dict) else {}
    fields = fields if isinstance(fields, dict) else {}

    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = [
        ('Date export', payload.get('exported_at', '')),
        ('UUID', payload.get('uuid', fields.get('uuid', ''))),
        ('Formulaire', payload.get('form_id', '')),
    ]
    for key in sorted(fields):
        if key in JSON_EXPORT_COLUMNS:
            continue
        summary_rows.append((humanize_export_field(key), fields.get(key)))
    _append_table(_new_sheet(wb, 'Synthese'), ['Champ', 'Valeur'], summary_rows)

    metiers = _safe_json(fields.get('metiers_json'), [])
    if isinstance(metiers, list):
        rows = [
            (
                item.get('metier', ''),
                item.get('mode', ''),
                item.get('etp', ''),
                item.get('etpCdi', ''),
                item.get('etpCdd', ''),
            )
            for item in metiers
            if isinstance(item, dict)
        ]
        _append_table(_new_sheet(wb, 'Metiers'), ['Métier', 'Mode', 'ETP', 'ETP CDI', 'ETP CDD'], rows)

    formations = _safe_json(fields.get('catalogue_formations_json'), [])
    if isinstance(formations, list):
        rows = [
            (
                item.get('nom', ''),
                item.get('niveau', ''),
                item.get('secteur', ''),
                item.get('nb', ''),
            )
            for item in formations
            if isinstance(item, dict)
        ]
        _append_table(_new_sheet(wb, 'Formations'), ['Formation', 'Niveau', 'Secteur', 'Nombre'], rows)

    prestations = _safe_json(fields.get('prestations_json'), {})
    prestation_rows = []
    if isinstance(prestations, dict):
        for section_key, section in prestations.items():
            if not isinstance(section, dict):
                continue
            label = _format_conditional_display_name(section_key)
            prestation_rows.append((label, humanize_export_path('done'), section.get('done', '')))
            for path, value in _flatten_for_xlsx(section):
                if path in {'done'} or should_skip_readable_export_path(path):
                    continue
                prestation_rows.append((label, humanize_export_path(path), value))
    _append_table(_new_sheet(wb, 'Prestations'), ['Section', 'Champ', 'Valeur'], prestation_rows)

    details = _details_payload(fields)
    wizard = details.get(WIZARD_STATE_KEY) if isinstance(details, dict) else {}
    runtime = wizard.get('runtime') if isinstance(wizard, dict) else {}
    conditional_defs = runtime.get('conditionalDefs') if isinstance(runtime, dict) else []
    conditional_state = runtime.get('conditionalState') if isinstance(runtime, dict) else {}
    conditional_rows = []
    if isinstance(conditional_defs, list) and isinstance(conditional_state, dict):
        for item in conditional_defs:
            if not isinstance(item, dict):
                continue
            cond_id = str(item.get('id') or '')
            name = _format_conditional_display_name(item.get('name') or cond_id)
            state = conditional_state.get(cond_id, {})
            if not isinstance(state, dict):
                continue
            conditional_rows.append((name, humanize_export_path('done'), state.get('done', '')))
            for path, value in _flatten_for_xlsx(state):
                if path == 'done' or should_skip_readable_export_path(path):
                    continue
                conditional_rows.append((name, humanize_export_path(path), value))
    _append_table(_new_sheet(wb, 'Etapes conditionnelles'), ['Étape', 'Champ', 'Valeur'], conditional_rows)

    selection_rows = []
    for column in ['prestations_orp_json', 'prestations_indirectes_json', 'finess_json']:
        parsed = _safe_json(fields.get(column), {})
        for path, value in _flatten_for_xlsx(parsed, column):
            selection_rows.append((humanize_export_path(path), value))
    _append_table(_new_sheet(wb, 'Selections'), ['Champ', 'Valeur'], selection_rows)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def compute_quick_step_progress(fields: dict) -> dict:
    """
    Fast progress estimator by major sections.
    Returns counts + list of done/remaining step labels.
    """
    finess_values = extract_finess_values(fields)
    identification_done = (
        _has_value(fields.get('es_nom'))
        and _has_value(fields.get('validateur_email'))
        and _has_value(fields.get('es_departement'))
        and len(finess_values) > 0
    )

    selected_dispositifs = [
        _as_bool(fields.get('check_esrp')),
        _as_bool(fields.get('check_espo')),
        _as_bool(fields.get('check_ueros')),
        _as_bool(fields.get('check_deac')),
    ]
    rh_done = any(selected_dispositifs) and _has_value(fields.get('metiers_json'))

    details_json = _details_payload(fields)
    prestations_orp_legacy = _safe_json(fields.get('prestations_orp_json'), {})
    prestations_orp_canonical = _details_get(details_json, ('prestations', 'selection', 'orp'), {})
    prestations_orp = {}
    if isinstance(prestations_orp_canonical, dict):
        prestations_orp.update(prestations_orp_canonical)
    if isinstance(prestations_orp_legacy, dict):
        prestations_orp.update(prestations_orp_legacy)
    vos_prestations_done = _has_value(fields.get('pec')) or _has_any_checked(prestations_orp)

    contexte_done = (
        _has_value(fields.get('q32_implantation'))
        and _has_value(fields.get('q33_transports'))
        and _has_value(fields.get('q53_afpa'))
    )

    steps = [
        ('Identification', identification_done),
        ('Volet RH', rh_done),
        ('Vos prestations', vos_prestations_done),
        ('Contexte écologique', contexte_done),
    ]

    # Conditional quick estimator.
    cond_expected_labels = []
    cond_done_labels = []
    wizard_runtime = _details_get(details_json, (WIZARD_STATE_KEY, 'runtime'), {})

    if isinstance(wizard_runtime, dict):
        runtime_defs = wizard_runtime.get('conditionalDefs')
        runtime_state = wizard_runtime.get('conditionalState')
        if isinstance(runtime_defs, list) and isinstance(runtime_state, dict):
            for item in runtime_defs:
                if not isinstance(item, dict):
                    continue
                cond_id = str(item.get('id') or '').strip()
                cond_name = _format_conditional_display_name(str(item.get('name') or '').strip())
                if not cond_id or not cond_name:
                    continue
                cond_expected_labels.append(cond_name)
                state = runtime_state.get(cond_id)
                if isinstance(state, dict) and (_as_bool(state.get('done')) or _as_bool(state.get('__completed'))):
                    cond_done_labels.append(cond_name)

    if not cond_expected_labels:
        if _as_bool(fields.get('check_esrp')):
            cond_expected_labels.append('Prestations totales ESRP')
        if _as_bool(fields.get('check_espo')):
            cond_expected_labels.append('Prestations totales ESPO')
        if _as_bool(fields.get('check_ueros')):
            cond_expected_labels.append('Prestations totales UEROS')
        if _as_bool(fields.get('check_deac')):
            cond_expected_labels.append('Prestations totales DEAc')

        prestations_json = _safe_json(fields.get('prestations_json'), {})
        prestations_json_canonical = _details_get(details_json, ('prestations', 'conditional', 'state_by_key'), {})
        if (not isinstance(prestations_json, dict) or not prestations_json) and isinstance(prestations_json_canonical, dict):
            prestations_json = prestations_json_canonical
        if isinstance(prestations_json, dict):
            if _has_any_number_data(prestations_json.get('esrp')):
                cond_done_labels.append('Prestations totales ESRP')
            if _has_any_number_data(prestations_json.get('espo')):
                cond_done_labels.append('Prestations totales ESPO')
            if _has_any_number_data(prestations_json.get('ueros')):
                cond_done_labels.append('Prestations totales UEROS')
            if _has_any_number_data(prestations_json.get('deac')):
                cond_done_labels.append('Prestations totales DEAc')

        orp_map = {
            'orp_pec': 'Prestation ORP: PEC',
            'orp_autre_eval': "Prestation ORP: Autre dispositif d'évaluation",
            'orp_orientation_espo': 'Prestation ORP: Orientation ESPO',
            'orp_parcours_sociopro': 'Prestation ORP: Parcours socio-professionnel',
            'orp_autre_parcours': 'Prestation ORP: Autre type de parcours',
            'orp_parcours_qualif': 'Prestation ORP: Parcours certifiant',
            'orp_remise_niveau': 'Prestation ORP: Remise à niveau',
            'orp_formation_pro': 'Prestation ORP: Formation professionnalisante',
            'orp_formation_certif': 'Prestation ORP: Formation certifiante',
            'orp_formation_accomp_pro': 'Prestation ORP: Formation accompagnée pro',
            'orp_formation_accomp_certif': 'Prestation ORP: Formation accompagnée certifiante',
        }
        selected_orp_keys = []
        if isinstance(prestations_orp, dict):
            for key, label in orp_map.items():
                if _as_bool(prestations_orp.get(key)):
                    selected_orp_keys.append((key, label))
                    cond_expected_labels.append(label)

        if isinstance(details_json, dict):
            for key, label in selected_orp_keys:
                state = details_json.get(key, {})
                if not isinstance(state, dict):
                    state = {}
                # New canonical location (schema v2)
                if not state:
                    canonical_state = _details_get(details_json, ('prestations', 'conditional', 'state_by_key', key), {})
                    if isinstance(canonical_state, dict):
                        state = canonical_state
                if isinstance(state, dict) and _as_bool(state.get('__completed')):
                    cond_done_labels.append(label)

    done_main = sum(1 for _, ok in steps if ok)
    total_main = len(steps)
    done_main_labels = [label for label, ok in steps if ok]
    remaining_main = [label for label, ok in steps if not ok]

    cond_expected_labels = list(dict.fromkeys(cond_expected_labels))
    cond_done_labels = list(dict.fromkeys(cond_done_labels))
    remaining_cond = [label for label in cond_expected_labels if label not in cond_done_labels]

    return {
        'main': {
            'done': done_main,
            'total': total_main,
            'done_labels': done_main_labels,
            'remaining_labels': remaining_main,
        },
        'conditional': {
            'done': len(cond_done_labels),
            'total': len(cond_expected_labels),
            'done_labels': cond_done_labels,
            'remaining_labels': remaining_cond,
        },
    }


def fetch_all_records(config: dict, headers: dict):
    """Fetch records for a form table (single pass, up to 5000 rows)."""
    url = f"{GRIST_BASE_URL}/api/docs/{config['doc_id']}/tables/{config['table_id']}/records"
    resp = requests.get(url, params={'limit': 5000}, headers=headers)
    if resp.status_code != 200:
        raise RuntimeError(f'Failed to read records: HTTP {resp.status_code} - {resp.text}')
    payload = resp.json()
    return payload.get('records', [])


def _parse_response_json_safe(resp):
    """Return JSON payload when possible, otherwise a readable fallback dict."""
    try:
        return resp.json()
    except Exception:
        body = (resp.text or '').strip()
        return {
            'error': f'Upstream response is not valid JSON (HTTP {resp.status_code}).',
            'raw': body[:500],
        }


def fetch_record_by_field(base_url: str, field_name: str, value: str, headers: dict):
    """Fetch a single record by a unique field from Grist."""
    filter_param = json.dumps({field_name: [value]})
    resp = requests.get(f"{base_url}/records", params={'filter': filter_param}, headers=headers)
    if resp.status_code != 200:
        return None, resp

    payload = _parse_response_json_safe(resp)
    if not isinstance(payload, dict) or not isinstance(payload.get('records'), list):
        raise RuntimeError('Failed to decode Grist record lookup response.')
    return payload['records'][0] if payload['records'] else None, resp


def resolve_record_key(allowed_columns: set[str]) -> str | None:
    """Pick the unique key used to upsert records for a target table."""
    if 'uuid' in allowed_columns:
        return 'uuid'
    if 'id_tally' in allowed_columns:
        return 'id_tally'
    return None


def _grist_write_redirect_response(resp):
    """Return a readable error when an upstream write endpoint redirects."""
    if not 300 <= resp.status_code < 400:
        return None

    source = str(getattr(resp, 'url', '') or '').strip()
    location = str(resp.headers.get('Location') or '').strip()
    source_part = f' Source URL: {source}.' if source else ''
    suffix = f' Redirect target: {location}' if location else ''
    return jsonify({
        'error': (
            f'Grist write endpoint redirected with HTTP {resp.status_code}. '
            f'Check GRIST_BASE_URL in production.{source_part}{suffix}'
        ),
    }), 502


def write_grist_records(method: str, url: str, payload: dict, headers: dict, max_redirects: int = 3):
    """Follow Grist write redirects while preserving the HTTP method and cookies."""
    current_url = url
    with requests.Session() as session:
        for _ in range(max_redirects + 1):
            resp = session.request(
                method,
                current_url,
                json=payload,
                headers=headers,
                allow_redirects=False,
            )
            if not 300 <= resp.status_code < 400:
                return resp

            location = str(resp.headers.get('Location') or '').strip()
            if not location:
                return resp
            current_url = urljoin(current_url, location)

    return resp


def fetch_table_records(doc_id: str, table_id: str, headers: dict, limit: int = 5000):
    """Read records from a Grist table."""
    url = f"{GRIST_BASE_URL}/api/docs/{doc_id}/tables/{table_id}/records"
    resp = requests.get(url, params={'limit': limit}, headers=headers)
    if resp.status_code != 200:
        raise RuntimeError(f'Failed to read {table_id}: HTTP {resp.status_code} - {resp.text}')
    payload = _parse_response_json_safe(resp)
    return payload.get('records', []) if isinstance(payload, dict) else []


def _month_key(value) -> str | None:
    """Extract YYYY-MM from common timestamp/date values."""
    txt = str(value or '').strip()
    if len(txt) >= 7 and txt[4] == '-':
        return txt[:7]
    for fmt in (
        '%Y-%m-%dT%H:%M:%S.%fZ',
        '%Y-%m-%dT%H:%M:%SZ',
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%d/%m/%Y',
    ):
        try:
            return datetime.strptime(txt, fmt).strftime('%Y-%m')
        except Exception:
            continue
    return None


def _parse_datetime(value) -> datetime | None:
    txt = str(value or '').strip()
    if not txt:
        return None
    try:
        return datetime.fromisoformat(txt.replace('Z', '+00:00'))
    except Exception:
        pass
    for fmt in (
        '%Y-%m-%dT%H:%M:%S.%fZ',
        '%Y-%m-%dT%H:%M:%SZ',
        '%Y-%m-%d',
    ):
        try:
            return datetime.strptime(txt, fmt)
        except Exception:
            continue
    return None


def _duration_hours(start_value, end_value) -> float | None:
    start = _parse_datetime(start_value)
    end = _parse_datetime(end_value)
    if not start or not end:
        return None
    delta_seconds = (end - start).total_seconds()
    if delta_seconds < 0:
        return None
    return delta_seconds / 3600.0


def _duration_summary(values: list[float]) -> dict:
    ordered = sorted(v for v in values if v is not None)
    if not ordered:
        return {'count': 0, 'avg_hours': None, 'median_hours': None, 'min_hours': None, 'max_hours': None}
    count = len(ordered)
    middle = count // 2
    median = ordered[middle] if count % 2 else (ordered[middle - 1] + ordered[middle]) / 2.0
    return {
        'count': count,
        'avg_hours': round(sum(ordered) / count, 1),
        'median_hours': round(median, 1),
        'min_hours': round(ordered[0], 1),
        'max_hours': round(ordered[-1], 1),
    }


def _split_multi_value(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = str(value).split('|')
    return [str(item).strip() for item in raw_items if str(item).strip()]


def _clean_public_label(label: str) -> str:
    return ' '.join(str(label or '').strip().split())


def _public_breakdown_label(kind: str, label: str) -> str:
    """Normalize raw stored values into cleaner public labels."""
    cleaned = _clean_public_label(label)
    lowered = cleaned.lower()

    if kind == 'matchings_par_statut':
        return {
            'calcule': 'Calcule',
            'a_valider_admin': 'A valider admin',
            'refuse_admin': 'Refuse admin',
            'valide_admin': 'Valide admin',
            'envoye_employeur': 'Envoye employeur',
            'accepte_employeur': 'Accepte employeur',
            'refuse_employeur': 'Refuse employeur',
            'mise_en_relation_faite': 'Mise en relation faite',
            'embauche_confirmee': 'Embauche confirmee',
        }.get(lowered, cleaned[:1].upper() + cleaned[1:])

    if kind == 'mobilite_candidats':
        if ':' in cleaned:
            prefix, raw_value = [part.strip() for part in cleaned.split(':', 1)]
            prefix_l = prefix.lower()
            if prefix_l == 'type':
                return raw_value
            if prefix_l in {'pays souhaités', 'pays souhaites'}:
                return f'Pays vises : {raw_value}'
        return cleaned[:1].upper() + cleaned[1:]

    if kind == 'experience_internationale_candidats':
        if ':' in cleaned:
            prefix, raw_value = [part.strip() for part in cleaned.split(':', 1)]
            prefix_l = prefix.lower()
            if prefix_l in {'expérience pays', 'experience pays'}:
                return f'Experience dans : {raw_value}'
        return cleaned[:1].upper() + cleaned[1:]

    if kind == 'retours_employeurs':
        return {
            'contact': 'Accepté par employeur',
            'not_contact': 'Refusé par employeur',
            'no_response': 'Sans réponse',
        }.get(lowered, cleaned[:1].upper() + cleaned[1:])

    return cleaned


def _counter_to_rows(counter: Counter) -> list[dict]:
    """Convert a counter to public rows without collapsing categories."""
    return [
        {'label': label, 'count': count}
        for label, count in sorted(counter.items(), key=lambda item: (-item[1], str(item[0]).lower()))
    ]


def _add_public_counter(counter: Counter, kind: str, raw_label: str):
    public_label = _public_breakdown_label(kind, raw_label)
    if public_label:
        counter[public_label] += 1


def _eures_public_sector_labels(value) -> list[str]:
    sectors = eures_candidate_sectors(str(value or ''))
    return [
        EURES_PUBLIC_SECTOR_LABELS[sector]
        for sector in EURES_PUBLIC_SECTOR_LABELS
        if sector in sectors
    ]


def _eures_public_mobility_labels(value) -> list[str]:
    labels: list[str] = []
    seen: set[str] = set()
    for part in _split_multi_value(value):
        if ':' in part:
            prefix, raw_values = [chunk.strip() for chunk in part.split(':', 1)]
            values = [item.strip() for item in raw_values.split(',') if item.strip()]
            if not values:
                values = [raw_values.strip()] if raw_values.strip() else []
            for item in values:
                if prefix.lower() == 'type':
                    label = item
                elif prefix.lower() in {'pays souhaités', 'pays souhaites'}:
                    label = f'Pays vises : {item}'
                else:
                    continue
                if label and label not in seen:
                    seen.add(label)
                    labels.append(label)
        else:
            for item in [chunk.strip() for chunk in str(part).split(',') if chunk.strip()]:
                if item and item not in seen:
                    seen.add(item)
                    labels.append(item)
    return labels


def _eures_public_experience_labels(value) -> list[str]:
    labels: list[str] = []
    seen: set[str] = set()
    for part in _split_multi_value(value):
        if ':' not in part:
            continue
        prefix, raw_values = [chunk.strip() for chunk in part.split(':', 1)]
        if prefix.lower() not in {'expérience pays', 'experience pays'}:
            continue
        values = [item.strip() for item in raw_values.split(',') if item.strip()]
        if not values:
            values = [raw_values.strip()] if raw_values.strip() else []
        for item in values:
            label = f'Experience dans : {item}'
            if label and label not in seen:
                seen.add(label)
                labels.append(label)
    return labels


def _safe_int(value) -> int:
    try:
        return int(float(value or 0))
    except Exception:
        return 0


def build_eures_public_stats() -> dict:
    """Build a public, non-nominative stats payload for EURES beta."""
    candidate_config = get_form_config('eures-beta', 'candidate')
    employer_config = get_form_config('eures-beta', 'employer')
    if not candidate_config or not employer_config:
        raise RuntimeError('EURES beta candidate/employer Grist configuration is incomplete.')

    candidate_headers = {'Accept': 'application/json'}
    if candidate_config.get('api_key'):
        candidate_headers['Authorization'] = f"Bearer {candidate_config['api_key']}"

    employer_headers = {'Accept': 'application/json'}
    if employer_config.get('api_key'):
        employer_headers['Authorization'] = f"Bearer {employer_config['api_key']}"

    candidats = fetch_table_records(candidate_config['doc_id'], EURES_CANDIDATS_TABLE, candidate_headers)
    besoins = fetch_table_records(employer_config['doc_id'], EURES_BESOINS_TABLE, employer_headers)
    matchings = fetch_table_records(candidate_config['doc_id'], EURES_MATCHINGS_TABLE, candidate_headers)

    monthly: dict[str, dict[str, int | str]] = defaultdict(lambda: {
        'mois': '',
        'candidats': 0,
        'besoins_employeurs': 0,
        'matchings': 0,
        'candidats_contactes': 0,
        'candidatures_transmises_employeur': 0,
        'contacts_acceptes_employeur': 0,
        'contacts_refuses_employeur': 0,
        'contacts_sans_reponse_employeur': 0,
        'embauches': 0,
    })

    besoins_par_pays = Counter()
    secteurs = Counter()
    mobilite_candidats = Counter()
    experience_internationale_candidats = Counter()
    matchings_par_statut = Counter()
    retours_employeurs = Counter()
    delays = {
        'calcul_to_admin': [],
        'admin_to_send': [],
        'send_to_employer_response': [],
        'response_to_relation': [],
        'relation_to_hire': [],
    }

    for rec in candidats:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        month = _month_key(fields.get('tally_submitted_at'))
        if month:
            monthly[month]['mois'] = month
            monthly[month]['candidats'] += 1
        for label in _eures_public_sector_labels(fields.get('metier')):
            _add_public_counter(secteurs, 'secteurs', label)
        for label in _eures_public_mobility_labels(fields.get('mobilite')):
            _add_public_counter(mobilite_candidats, 'mobilite_candidats', label)
        for label in _eures_public_experience_labels(fields.get('mobilite')):
            _add_public_counter(experience_internationale_candidats, 'experience_internationale_candidats', label)

    for rec in besoins:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        month = _month_key(fields.get('tally_submitted_at'))
        if month:
            monthly[month]['mois'] = month
            monthly[month]['besoins_employeurs'] += 1
        for label in _split_multi_value(fields.get('pays_normalise')):
            _add_public_counter(besoins_par_pays, 'besoins_par_pays', label)
        for label in _eures_public_sector_labels(fields.get('poste')):
            _add_public_counter(secteurs, 'secteurs', label)

    for rec in matchings:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        month = _month_key(fields.get('date_calcul'))
        if month:
            monthly[month]['mois'] = month
            monthly[month]['matchings'] += 1
        workflow_status = _matching_workflow_status(fields)
        if workflow_status:
            _add_public_counter(matchings_par_statut, 'matchings_par_statut', workflow_status)

        sent_month = _month_key(fields.get('sent_to_employer_at'))
        if sent_month:
            monthly[sent_month]['mois'] = sent_month
            monthly[sent_month]['candidatures_transmises_employeur'] += 1

        relation_month = _month_key(fields.get('mise_en_relation_at'))
        if relation_month:
            monthly[relation_month]['mois'] = relation_month
            monthly[relation_month]['candidats_contactes'] += 1

        hire_month = _month_key(fields.get('embauche_confirmee_at'))
        if hire_month:
            monthly[hire_month]['mois'] = hire_month
            monthly[hire_month]['embauches'] += 1

        response_bucket = None
        response_month = None
        if workflow_status in {'accepte_employeur', 'mise_en_relation_faite', 'embauche_confirmee'}:
            response_bucket = 'contact'
            response_month = _month_key(fields.get('employer_response_at') or fields.get('mise_en_relation_at') or fields.get('embauche_confirmee_at'))
        elif workflow_status == 'refuse_employeur':
            response_bucket = 'not_contact'
            response_month = _month_key(fields.get('employer_response_at'))
        elif workflow_status == 'envoye_employeur':
            response_bucket = 'no_response'
            response_month = sent_month

        if response_bucket:
            _add_public_counter(retours_employeurs, 'retours_employeurs', response_bucket)
            if response_month:
                monthly[response_month]['mois'] = response_month
                if response_bucket == 'contact':
                    monthly[response_month]['contacts_acceptes_employeur'] += 1
                elif response_bucket == 'not_contact':
                    monthly[response_month]['contacts_refuses_employeur'] += 1
                else:
                    monthly[response_month]['contacts_sans_reponse_employeur'] += 1

        duration = _duration_hours(fields.get('date_calcul'), fields.get('admin_decision_at'))
        if duration is not None:
            delays['calcul_to_admin'].append(duration)
        duration = _duration_hours(fields.get('admin_decision_at'), fields.get('sent_to_employer_at'))
        if duration is not None:
            delays['admin_to_send'].append(duration)
        duration = _duration_hours(fields.get('sent_to_employer_at'), fields.get('employer_response_at'))
        if duration is not None:
            delays['send_to_employer_response'].append(duration)
        duration = _duration_hours(fields.get('employer_response_at'), fields.get('mise_en_relation_at'))
        if duration is not None:
            delays['response_to_relation'].append(duration)
        duration = _duration_hours(fields.get('mise_en_relation_at'), fields.get('embauche_confirmee_at'))
        if duration is not None:
            delays['relation_to_hire'].append(duration)

    manual_stats_available = False
    stats_config = get_eures_stats_config()
    if stats_config:
        stats_headers = {'Accept': 'application/json'}
        if stats_config.get('api_key'):
            stats_headers['Authorization'] = f"Bearer {stats_config['api_key']}"
        try:
            manual_rows = fetch_table_records(stats_config['doc_id'], stats_config['table_id'], stats_headers)
            manual_stats_available = True
        except Exception:
            manual_rows = []
        for rec in manual_rows:
            fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
            month = _month_key(fields.get('mois'))
            if not month:
                continue
            monthly[month]['mois'] = month
            monthly[month]['candidats_contactes'] += 0
            monthly[month]['candidatures_transmises_employeur'] += 0
            monthly[month]['embauches'] += 0

    monthly_rows = [row for _, row in sorted(monthly.items()) if row.get('mois')]

    totals = {
        'candidats': len(candidats),
        'besoins_employeurs': len(besoins),
        'matchings': len(matchings),
        'candidats_contactes': sum(int(row['candidats_contactes']) for row in monthly_rows),
        'candidatures_transmises_employeur': sum(int(row['candidatures_transmises_employeur']) for row in monthly_rows),
        'contacts_acceptes_employeur': sum(int(row['contacts_acceptes_employeur']) for row in monthly_rows),
        'contacts_refuses_employeur': sum(int(row['contacts_refuses_employeur']) for row in monthly_rows),
        'contacts_sans_reponse_employeur': sum(int(row['contacts_sans_reponse_employeur']) for row in monthly_rows),
        'embauches': sum(int(row['embauches']) for row in monthly_rows),
    }

    return {
        'ok': True,
        'generated_at': time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),
        'manual_stats_table': {
            'configured': manual_stats_available,
            'table_id': stats_config['table_id'] if stats_config else None,
        },
        'totals': totals,
        'durations': {
            'calcul_to_admin': _duration_summary(delays['calcul_to_admin']),
            'admin_to_send': _duration_summary(delays['admin_to_send']),
            'send_to_employer_response': _duration_summary(delays['send_to_employer_response']),
            'response_to_relation': _duration_summary(delays['response_to_relation']),
            'relation_to_hire': _duration_summary(delays['relation_to_hire']),
        },
        'monthly': monthly_rows,
        'breakdowns': {
            'besoins_par_pays': _counter_to_rows(besoins_par_pays),
            'secteurs': _counter_to_rows(secteurs),
            'mobilite_candidats': _counter_to_rows(mobilite_candidats),
            'experience_internationale_candidats': _counter_to_rows(experience_internationale_candidats),
            'matchings_par_statut': _counter_to_rows(matchings_par_statut),
            'retours_employeurs': _counter_to_rows(retours_employeurs),
        },
    }


def fetch_matching_record(doc_id: str, besoin_id: str, candidat_id: str, headers: dict):
    """Read one matching by besoin_id/candidat_id pair."""
    url = f"{GRIST_BASE_URL}/api/docs/{doc_id}/tables/{EURES_MATCHINGS_TABLE}/records"
    filter_param = json.dumps({'besoin_id': [besoin_id], 'candidat_id': [candidat_id]})
    resp = requests.get(url, params={'filter': filter_param}, headers=headers)
    if resp.status_code != 200:
        raise RuntimeError(f'Failed to read Matchings: HTTP {resp.status_code} - {resp.text}')
    payload = _parse_response_json_safe(resp)
    records = payload.get('records', []) if isinstance(payload, dict) else []
    return records[0] if records else None


def upsert_matching_record(doc_id: str, fields: dict, headers: dict):
    """Create or update a matching record from besoin_id/candidat_id."""
    besoin_id = str(fields.get('besoin_id') or '').strip()
    candidat_id = str(fields.get('candidat_id') or '').strip()
    if not besoin_id or not candidat_id:
        raise RuntimeError('Missing besoin_id or candidat_id for matching upsert.')

    table_config = {'doc_id': doc_id, 'table_id': EURES_MATCHINGS_TABLE}
    ensure_table_columns(table_config, set(fields.keys()) & (EURES_MATCHING_FIELDS | EURES_MATCHING_ADMIN_FIELDS), headers)
    existing = fetch_matching_record(doc_id, besoin_id, candidat_id, headers)
    merged_fields = dict(fields)
    if existing and isinstance(existing, dict):
        existing_fields = existing.get('fields', {}) if isinstance(existing.get('fields'), dict) else {}
        for key in EURES_MATCHING_ADMIN_FIELDS:
            if key not in merged_fields and key in existing_fields:
                merged_fields[key] = existing_fields.get(key)
        if not merged_fields.get('workflow_status'):
            merged_fields['workflow_status'] = _matching_workflow_status(existing_fields)
    else:
        merged_fields.setdefault('workflow_status', _initial_matching_workflow_status(fields.get('statut', '')))

    ensure_table_columns(table_config, set(merged_fields.keys()) & (EURES_MATCHING_FIELDS | EURES_MATCHING_ADMIN_FIELDS), headers)
    allowed_columns = get_table_columns(table_config, headers)
    filtered_fields = {k: v for k, v in merged_fields.items() if k in allowed_columns}
    base_url = f"{GRIST_BASE_URL}/api/docs/{doc_id}/tables/{EURES_MATCHINGS_TABLE}/records"

    if existing:
        payload = {'records': [{'id': existing['id'], 'fields': filtered_fields}]}
        resp = write_grist_records('PATCH', base_url, payload, headers)
    else:
        payload = {'records': [{'fields': filtered_fields}]}
        resp = write_grist_records('POST', base_url, payload, headers)

    if resp.status_code != 200:
        raise RuntimeError(f'Failed to write Matchings: HTTP {resp.status_code} - {resp.text}')
    return resp


def update_matching_record_by_id(doc_id: str, record_id: int, fields: dict, headers: dict):
    """Patch a matching record by Grist record id."""
    table_config = {'doc_id': doc_id, 'table_id': EURES_MATCHINGS_TABLE}
    ensure_table_columns(table_config, set(fields.keys()) & (EURES_MATCHING_FIELDS | EURES_MATCHING_ADMIN_FIELDS), headers)
    allowed_columns = get_table_columns(table_config, headers)
    filtered_fields = {k: v for k, v in fields.items() if k in allowed_columns}
    base_url = f"{GRIST_BASE_URL}/api/docs/{doc_id}/tables/{EURES_MATCHINGS_TABLE}/records"
    payload = {'records': [{'id': record_id, 'fields': filtered_fields}]}
    resp = write_grist_records('PATCH', base_url, payload, headers)
    if resp.status_code != 200:
        raise RuntimeError(f'Failed to update Matchings: HTTP {resp.status_code} - {resp.text}')
    return resp


def fetch_record_by_id(doc_id: str, table_id: str, record_id: int, headers: dict):
    """Read one Grist record by numeric id."""
    url = f"{GRIST_BASE_URL}/api/docs/{doc_id}/tables/{table_id}/records"
    resp = requests.get(url, params={'filter': json.dumps({'id': [record_id]})}, headers=headers)
    if resp.status_code != 200:
        raise RuntimeError(f'Failed to read {table_id}: HTTP {resp.status_code} - {resp.text}')
    payload = _parse_response_json_safe(resp)
    records = payload.get('records', []) if isinstance(payload, dict) else []
    return records[0] if records else None


def eures_normalize_text(value) -> str:
    return str(value or '').strip().lower()


def eures_canonical_sector(value: str) -> str:
    value_l = eures_normalize_text(value)
    for token, sector in EURES_SECTOR_CANONICAL_MAP.items():
        if token in value_l:
            return sector
    return ''


def eures_candidate_sectors(value: str) -> set[str]:
    value_l = eures_normalize_text(value)
    return {sector for token, sector in EURES_SECTOR_CANONICAL_MAP.items() if token in value_l}


def eures_parse_language_requirements(value: str) -> dict[str, int]:
    levels = {
        'pas nécessaire': 0,
        'pas necessaire': 0,
        'not necessary': 0,
        'communication professionelle': 2,
        'communication professionnelle': 2,
        'professional communication': 2,
        'pouvoir communiquer dans la langue locale': 2,
        'très important au quotidien': 3,
        'tres important au quotidien': 3,
        'very important on a daily basis': 3,
    }
    result: dict[str, int] = {}
    for part in str(value or '').split('|'):
        if ':' not in part:
            continue
        lang, raw_level = part.split(':', 1)
        lang_key = eures_normalize_text(lang)
        level_text = eures_normalize_text(raw_level)
        score = next((v for k, v in levels.items() if k in level_text), None)
        if score is not None:
            result[lang_key] = score
    return result


def eures_parse_candidate_languages(value: str) -> dict[str, int]:
    levels = {
        'je ne la pratique pas': 0,
        'i do not speak it': 0,
        "j'utilise un outil de traduction si nécessaire": 0,
        "j'utilise un outil de traduction si necessaire": 0,
        'translation tool': 0,
        'je peux communiquer simplement': 2,
        "je peux échanger à l'oral": 2,
        "je peux echanger a l'oral": 2,
        'simple communication': 2,
        'je peux travailler avec cette langue': 3,
        'langue maternelle': 4,
        'native': 4,
    }
    result: dict[str, int] = {}
    for part in str(value or '').split('|'):
        if ':' not in part:
            continue
        lang, raw_level = part.split(':', 1)
        lang_key = eures_normalize_text(lang)
        level_text = eures_normalize_text(raw_level)
        matched = [v for k, v in levels.items() if k in level_text]
        if matched:
            result[lang_key] = max(matched)
    return result


def eures_availability_rank(value: str) -> int | None:
    value_l = eures_normalize_text(value)
    ordered = [
        ('dès que possible', 0),
        ('des que possible', 0),
        ('dans les prochains jours', 1),
        ('dans les prochaines semaines', 2),
        ('dans 1 à 3 mois', 3),
        ('dans 1 a 3 mois', 3),
        ('je ne sais pas encore', 4),
    ]
    for token, rank in ordered:
        if token in value_l:
            return rank
    return None


def eures_to_float(value):
    if value in (None, '', 0, '0'):
        return None
    try:
        return float(value)
    except Exception:
        return None


def eures_score_sector_fit(expected: str, actual: str) -> tuple[int, str]:
    expected_sector = eures_canonical_sector(expected)
    candidate_sectors = eures_candidate_sectors(actual)
    if not expected_sector:
        return 0, 'metier: secteur employeur absent'
    if not candidate_sectors:
        return 0, 'metier: secteur candidat absent'
    if expected_sector in candidate_sectors:
        if len(candidate_sectors) == 1:
            return 30, f'metier: secteur exact ({expected_sector})'
        return 20, f'metier: secteur present ({expected_sector})'
    return 0, f'metier: secteur non aligne ({expected_sector})'


def eures_score_languages(expected: str, actual: str) -> tuple[int, str]:
    expected_map = eures_parse_language_requirements(expected)
    actual_map = eures_parse_candidate_languages(actual)
    scored = {lang: lvl for lang, lvl in expected_map.items() if lvl > 0}
    if not scored or not actual_map:
        return 0, 'langues: information manquante'
    matched = []
    strong = 0
    partial = 0
    for lang, required_level in scored.items():
        candidate_level = actual_map.get(lang)
        if candidate_level is None:
            continue
        if candidate_level >= required_level:
            strong += 1
            matched.append(lang)
        elif candidate_level >= 2 and candidate_level + 1 >= required_level:
            partial += 1
            matched.append(f'{lang} (partiel)')
    ratio = (strong + 0.5 * partial) / max(len(scored), 1)
    points = round(25 * ratio)
    if points > 0:
        return points, 'langues: ' + ', '.join(matched)
    return 0, 'langues: aucune correspondance'


def eures_score_location(besoin_pays: str, candidat_pays: str, mobilite: str) -> tuple[int, str]:
    besoin_pays_l = eures_normalize_text(besoin_pays)
    candidat_pays_l = eures_normalize_text(candidat_pays)
    mobilite_l = eures_normalize_text(mobilite)
    if besoin_pays_l and candidat_pays_l and (besoin_pays_l == candidat_pays_l or besoin_pays_l in candidat_pays_l):
        return 15, 'pays/mobilite: meme pays'
    if besoin_pays_l and mobilite_l and besoin_pays_l in mobilite_l:
        return 15, 'pays/mobilite: pays souhaite explicite'
    if mobilite_l and any(token in mobilite_l for token in ('transfrontali', 'expatriation', 'pays souhaités', 'pays souhaites')):
        return 10, 'pays/mobilite: mobilite declaree'
    return 0, 'pays/mobilite: contrainte geographique'


def eures_score_availability(date_debut: str, disponibilite: str) -> tuple[int, str]:
    disponibilite_l = str(disponibilite or '').strip()
    if not disponibilite_l:
        return 0, 'disponibilite: information manquante'
    besoin_rank = eures_availability_rank(date_debut)
    candidat_rank = eures_availability_rank(disponibilite)
    if besoin_rank is not None and candidat_rank is not None:
        if candidat_rank <= besoin_rank:
            return 15, 'disponibilite: date compatible'
        if candidat_rank == besoin_rank + 1:
            return 9, 'disponibilite: leger decalage'
        return 0, 'disponibilite: delai trop long'
    if eures_normalize_text(date_debut) and eures_normalize_text(date_debut) in eures_normalize_text(disponibilite):
        return 15, 'disponibilite: date compatible'
    return 7, 'disponibilite: verification manuelle recommandee'


def eures_candidate_salary_expectation(secteur: str, fields: dict):
    salary_fields = EURES_CANDIDAT_SECTOR_SALARY_FIELDS.get(secteur)
    if not salary_fields:
        return None
    type_field, min_field = salary_fields
    salary_type = str(fields.get(type_field) or '').strip()
    salary_min = eures_to_float(fields.get(min_field))
    if salary_min is None:
        return None
    return salary_type, salary_min


def eures_employer_salary_offer(secteur: str, fields: dict):
    salary_fields = EURES_EMPLOYEUR_SECTOR_SALARY_FIELDS.get(secteur)
    if not salary_fields:
        return None
    type_field, min_field, max_field = salary_fields
    salary_type = str(fields.get(type_field) or '').strip()
    salary_min = eures_to_float(fields.get(min_field))
    salary_max = eures_to_float(fields.get(max_field))
    if salary_min is None and salary_max is None:
        return None
    return salary_type, salary_min, salary_max


def eures_score_salary(secteur: str, candidat_fields: dict, besoin_fields: dict) -> tuple[int, str]:
    if not secteur:
        return 0, 'salaire: secteur indetermine'
    candidat_salary = eures_candidate_salary_expectation(secteur, candidat_fields)
    employeur_salary = eures_employer_salary_offer(secteur, besoin_fields)
    if candidat_salary is None or employeur_salary is None:
        return 6, 'salaire: donnees absentes, verification manuelle'
    _, candidat_min = candidat_salary
    _, employeur_min, employeur_max = employeur_salary
    if employeur_max and candidat_min <= employeur_max:
        if employeur_min and candidat_min < employeur_min:
            return 12, 'salaire: attente sous la fourchette proposee'
        return 15, 'salaire: compatible'
    if employeur_min and candidat_min <= employeur_min * 1.1:
        return 8, 'salaire: ecart limite'
    return 0, 'salaire: attente superieure au salaire propose'


def eures_matching_status(score: int) -> str:
    if score >= 75:
        return 'auto_envoyable'
    if score >= 55:
        return 'a_valider'
    return 'a_ecarter'


def compute_eures_matching(besoin_fields: dict, candidat_fields: dict) -> dict:
    """Compute a matching payload for one candidate/employer pair."""
    besoin_country = besoin_fields.get('pays_normalise') or besoin_fields.get('pays') or ''
    candidat_country = candidat_fields.get('pays_normalise') or candidat_fields.get('pays') or ''
    secteur = eures_canonical_sector(str(besoin_fields.get('poste') or ''))

    score_metier, raison_metier = eures_score_sector_fit(
        str(besoin_fields.get('poste') or ''),
        str(candidat_fields.get('metier') or ''),
    )
    score_langues, raison_langues = eures_score_languages(
        str(besoin_fields.get('langues_requises') or ''),
        str(candidat_fields.get('langues') or ''),
    )
    score_mobilite, raison_mobilite = eures_score_location(
        str(besoin_country),
        str(candidat_country),
        str(candidat_fields.get('mobilite') or ''),
    )
    score_disponibilite, raison_disponibilite = eures_score_availability(
        str(besoin_fields.get('date_debut') or ''),
        str(candidat_fields.get('disponibilite') or ''),
    )
    score_salaire, raison_salaire = eures_score_salary(secteur, candidat_fields, besoin_fields)

    reasons = []
    weaknesses = []
    for points, text in [
        (score_metier, raison_metier),
        (score_langues, raison_langues),
        (score_mobilite, raison_mobilite),
        (score_disponibilite, raison_disponibilite),
        (score_salaire, raison_salaire),
    ]:
        (reasons if points else weaknesses).append(text)

    score = score_metier + score_langues + score_mobilite + score_disponibilite + score_salaire
    return {
        'score': score,
        'score_metier': score_metier,
        'score_langues': score_langues,
        'score_mobilite': score_mobilite,
        'score_disponibilite': score_disponibilite,
        'score_salaire': score_salaire,
        'statut': eures_matching_status(score),
        'raisons': ' | '.join(reasons),
        'points_faibles': ' | '.join(weaknesses),
        'date_calcul': time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),
    }


def run_eures_matching_for_saved_record(form_id: str, role: str, saved_record: dict, config: dict, headers: dict):
    """Compute matchings for one newly saved EURES beta record and write them to Matchings."""
    if form_id != 'eures-beta':
        return {'processed': False, 'reason': 'unsupported_form'}
    if role not in {'candidate', 'employer'}:
        return {'processed': False, 'reason': 'unsupported_role'}

    all_candidats = fetch_table_records(config['doc_id'], EURES_CANDIDATS_TABLE, headers)
    all_besoins = fetch_table_records(config['doc_id'], EURES_BESOINS_TABLE, headers)
    saved_fields = saved_record.get('fields', {}) if isinstance(saved_record.get('fields'), dict) else {}

    writes = 0
    if role == 'candidate':
        for besoin in all_besoins:
            besoin_fields = besoin.get('fields', {}) if isinstance(besoin.get('fields'), dict) else {}
            candidat_id = str(saved_fields.get('id_tally') or saved_fields.get('uuid') or '')
            besoin_id = str(besoin_fields.get('id_tally') or besoin_fields.get('uuid') or '')
            if not candidat_id or not besoin_id:
                continue
            matching = compute_eures_matching(besoin_fields, saved_fields)
            payload = {'besoin_id': besoin_id, 'candidat_id': candidat_id, **matching}
            upsert_matching_record(config['doc_id'], payload, headers)
            writes += 1
    else:
        for candidat in all_candidats:
            candidat_fields = candidat.get('fields', {}) if isinstance(candidat.get('fields'), dict) else {}
            candidat_id = str(candidat_fields.get('id_tally') or candidat_fields.get('uuid') or '')
            besoin_id = str(saved_fields.get('id_tally') or saved_fields.get('uuid') or '')
            if not candidat_id or not besoin_id:
                continue
            matching = compute_eures_matching(saved_fields, candidat_fields)
            payload = {'besoin_id': besoin_id, 'candidat_id': candidat_id, **matching}
            upsert_matching_record(config['doc_id'], payload, headers)
            writes += 1

    return {'processed': True, 'writes': writes, 'role': role}


def _eures_admin_headers(config: dict) -> dict:
    headers = {'Accept': 'application/json'}
    if config.get('api_key'):
        headers['Authorization'] = f"Bearer {config['api_key']}"
    return headers


def _eures_admin_status(fields: dict) -> str:
    status = str((fields or {}).get('admin_status') or '').strip().lower()
    if status in {'accepted', 'refused', 'pending'}:
        return status
    return 'pending'


def _eures_workflow_transition_allowed(current_status: str, target_status: str) -> bool:
    allowed = {
        'calcule': {'a_valider_admin', 'refuse_admin'},
        'a_valider_admin': {'valide_admin', 'refuse_admin'},
        'valide_admin': {'envoye_employeur', 'refuse_admin'},
        'envoye_employeur': {'accepte_employeur', 'refuse_employeur', 'mise_en_relation_faite'},
        'accepte_employeur': {'mise_en_relation_faite', 'embauche_confirmee'},
        'mise_en_relation_faite': {'embauche_confirmee'},
        'refuse_employeur': set(),
        'refuse_admin': set(),
        'embauche_confirmee': set(),
    }
    if current_status == target_status:
        return True
    return target_status in allowed.get(current_status, set())


def _split_matching_text(value: str) -> list[str]:
    return [part.strip() for part in str(value or '').split('|') if part.strip()]


def _extract_employer_contact_email(contact_value: str) -> str:
    raw = str(contact_value or '').strip()
    if '@' not in raw:
        return ''
    if '|' in raw:
        for part in [p.strip() for p in raw.split('|') if p.strip()]:
            if '@' in part:
                return part
    return raw


def _extract_first_email(value) -> str:
    raw = str(value or '').strip()
    if not raw or '@' not in raw:
        return ''
    match = re.search(r'([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})', raw, flags=re.IGNORECASE)
    return normalize_email(match.group(1)) if match else ''


def _resolve_employer_recipient(employeur: dict) -> str:
    if not isinstance(employeur, dict):
        return ''

    direct_candidates = [
        employeur.get('contact', ''),
        employeur.get('email', ''),
        employeur.get('mail', ''),
        employeur.get('contact_email', ''),
        employeur.get('email_contact', ''),
    ]
    for value in direct_candidates:
        email = _extract_employer_contact_email(value) or _extract_first_email(value)
        if email:
            return email

    for value in employeur.values():
        email = _extract_first_email(value)
        if email:
            return email
    return ''


def _format_email_multiline(value: str) -> str:
    text = str(value or '').strip()
    if not text:
        return 'Non renseigné'
    return '<br>'.join(escape(part.strip()) for part in text.split('|') if part.strip()) or escape(text)


def _build_eures_feedback_url(record_id: int, response: str) -> str:
    token = get_eures_email_action_serializer().dumps({
        'record_id': int(record_id),
        'response': response,
    })
    return f"{get_public_app_base_url()}/eures-beta/matching-feedback?token={token}"


def build_brevo_matching_email(row: dict) -> tuple[str, str, str, str]:
    """Build recipient, subject, text body and HTML body for one accepted matching."""
    candidat = row.get('candidat', {}) if isinstance(row.get('candidat'), dict) else {}
    employeur = row.get('employeur', {}) if isinstance(row.get('employeur'), dict) else {}
    recipient = _resolve_employer_recipient(employeur)
    if not recipient:
        raise RuntimeError('Employer contact email is missing for accepted matching.')

    poste = employeur.get('poste', '') or 'Poste non renseigné'
    employeur_name = employeur.get('employeur', '') or 'votre structure'
    subject = f"[EURES beta] Proposition de candidature pour votre besoin - {poste}"

    reasons = row.get('raisons', [])
    if not isinstance(reasons, list):
        reasons = _split_matching_text(reasons)

    reasons_text = '\n'.join(f"- {item}" for item in reasons) if reasons else "- Profil cohérent avec votre besoin"
    reasons_html = ''.join(
        f"<li style=\"margin:0 0 8px;\">{escape(item)}</li>" for item in reasons
    ) or "<li style=\"margin:0 0 8px;\">Profil cohérent avec votre besoin</li>"

    candidate_name = str(candidat.get('nom') or 'Profil candidat')
    candidate_phone = str(candidat.get('telephone') or 'Non renseigné')
    candidate_city = str(candidat.get('ville') or 'Non renseignée')
    signature_name = get_eures_mail_signature_name()
    contact_yes_url = _build_eures_feedback_url(int(row.get('record_id') or 0), 'contact')
    contact_no_url = _build_eures_feedback_url(int(row.get('record_id') or 0), 'not_contact')
    body_text = (
        f"Bonjour,\n\n"
        f"Nous vous proposons un profil susceptible de correspondre à votre besoin de recrutement pour le poste : {poste}.\n\n"
        f"Entreprise : {employeur_name}\n"
        f"Poste recherché : {poste}\n"
        f"Date de début souhaitée : {employeur.get('date_debut', 'Non renseignée')}\n\n"
        f"Candidat\n"
        f"- Nom : {candidate_name}\n"
        f"- Email : {candidat.get('email', 'Non renseigné')}\n"
        f"- Téléphone : {candidate_phone}\n"
        f"- Ville actuelle : {candidate_city}\n"
        f"- Pays de résidence : {candidat.get('pays', 'Non renseigné')}\n"
        f"- Métier / secteur : {candidat.get('metier', 'Non renseigné')}\n"
        f"- Compétences : {candidat.get('competences', 'Non renseigné')}\n"
        f"- Langues : {candidat.get('langues', 'Non renseigné')}\n"
        f"- Mobilité : {candidat.get('mobilite', 'Non renseigné')}\n"
        f"- Disponibilité : {candidat.get('disponibilite', 'Non renseigné')}\n\n"
        f"Pourquoi ce profil a été retenu\n"
        f"{reasons_text}\n\n"
        "Actions rapides\n"
        f"- Je vais le contacter : {contact_yes_url}\n"
        f"- Je ne vais pas le contacter : {contact_no_url}\n\n"
        "Vous pouvez aussi répondre directement à cet email afin que nous poursuivions la mise en relation.\n\n"
        "Cordialement,\n"
        f"{signature_name}\n"
        "EURES beta\n"
    )

    body_html = f"""
<!doctype html>
<html lang="fr">
  <body style="margin:0;padding:0;background:#f4efe6;font-family:Georgia,'Times New Roman',serif;color:#1f1f1f;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f4efe6;padding:24px 12px;">
      <tr>
        <td align="center">
          <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:720px;background:#fffdf9;border:1px solid #e7dcc7;border-radius:18px;overflow:hidden;">
            <tr>
              <td style="padding:28px 32px;background:linear-gradient(135deg,#103a2b 0%,#1f5a45 100%);color:#ffffff;">
                <div style="font-size:13px;letter-spacing:1.6px;text-transform:uppercase;opacity:0.82;">EURES beta</div>
                <h1 style="margin:10px 0 0;font-size:30px;line-height:1.2;font-weight:700;">Proposition de candidature</h1>
                <p style="margin:12px 0 0;font-size:16px;line-height:1.6;max-width:560px;">
                  Nous vous transmettons un profil pouvant correspondre à votre besoin de recrutement pour le poste de
                  <strong>{escape(poste)}</strong>.
                </p>
              </td>
            </tr>
            <tr>
              <td style="padding:28px 32px 12px;">
                <p style="margin:0 0 18px;font-size:16px;line-height:1.7;">
                  Bonjour,
                </p>
                <p style="margin:0 0 22px;font-size:16px;line-height:1.7;">
                  Après analyse de votre besoin, nous avons identifié un profil candidat qui nous semble pertinent pour
                  <strong>{escape(employeur_name)}</strong>.
                </p>
              </td>
            </tr>
            <tr>
              <td style="padding:0 32px 28px;">
                <table role="presentation" width="100%" cellspacing="0" cellpadding="0">
                  <tr>
                    <td valign="top" width="50%" style="padding:0 10px 0 0;">
                      <div style="height:100%;background:#f8f4ec;border:1px solid #eadfcb;border-radius:14px;padding:20px;">
                        <div style="font-size:12px;letter-spacing:1.3px;text-transform:uppercase;color:#6a5a42;margin-bottom:10px;">Votre besoin</div>
                        <h2 style="margin:0 0 14px;font-size:22px;line-height:1.3;color:#103a2b;">{escape(poste)}</h2>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Entreprise :</strong><br>{escape(employeur_name)}</p>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Lieu / pays :</strong><br>{escape(str(employeur.get('pays') or 'Non renseigné'))}</p>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Langues attendues :</strong><br>{_format_email_multiline(employeur.get('langues_requises', ''))}</p>
                        <p style="margin:0;font-size:15px;line-height:1.6;"><strong>Date de début :</strong><br>{escape(str(employeur.get('date_debut') or 'Non renseignée'))}</p>
                      </div>
                    </td>
                    <td valign="top" width="50%" style="padding:0 0 0 10px;">
                      <div style="height:100%;background:#fffaf2;border:1px solid #eadfcb;border-radius:14px;padding:20px;">
                        <div style="font-size:12px;letter-spacing:1.3px;text-transform:uppercase;color:#6a5a42;margin-bottom:10px;">Profil proposé</div>
                        <h2 style="margin:0 0 14px;font-size:22px;line-height:1.3;color:#103a2b;">{escape(candidate_name)}</h2>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Email :</strong><br>{escape(str(candidat.get('email') or 'Non renseigné'))}</p>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Téléphone :</strong><br>{escape(candidate_phone)}</p>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Ville actuelle :</strong><br>{escape(candidate_city)}</p>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Pays de résidence :</strong><br>{escape(str(candidat.get('pays') or 'Non renseigné'))}</p>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Métier / secteur :</strong><br>{_format_email_multiline(candidat.get('metier', ''))}</p>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Langues :</strong><br>{_format_email_multiline(candidat.get('langues', ''))}</p>
                        <p style="margin:0 0 10px;font-size:15px;line-height:1.6;"><strong>Mobilité :</strong><br>{_format_email_multiline(candidat.get('mobilite', ''))}</p>
                        <p style="margin:0;font-size:15px;line-height:1.6;"><strong>Disponibilité :</strong><br>{_format_email_multiline(candidat.get('disponibilite', ''))}</p>
                      </div>
                    </td>
                  </tr>
                </table>
              </td>
            </tr>
            <tr>
              <td style="padding:0 32px 12px;">
                <div style="background:#fbf7ef;border-left:4px solid #c6932d;border-radius:12px;padding:18px 20px;">
                  <div style="font-size:12px;letter-spacing:1.3px;text-transform:uppercase;color:#7a6232;margin-bottom:10px;">Pourquoi ce profil</div>
                  <ul style="margin:0;padding-left:20px;font-size:15px;line-height:1.7;">
                    {reasons_html}
                  </ul>
                </div>
              </td>
            </tr>
            <tr>
              <td style="padding:12px 32px 0;">
                <div style="background:#103a2b;border-radius:14px;padding:22px;color:#ffffff;">
                  <p style="margin:0 0 10px;font-size:17px;line-height:1.6;"><strong>Suite proposée</strong></p>
                  <p style="margin:0;font-size:15px;line-height:1.7;">
                    Si ce profil retient votre attention, vous pouvez le contacter directement ou nous répondre par email.
                  </p>
                  <table role="presentation" cellspacing="0" cellpadding="0" style="margin-top:18px;">
                    <tr>
                      <td style="padding:0 12px 12px 0;">
                        <a href="{escape(contact_yes_url)}" style="display:inline-block;padding:12px 18px;border-radius:999px;background:#f2c04c;color:#173a2a;text-decoration:none;font-size:14px;font-weight:700;">Je vais le contacter</a>
                      </td>
                      <td style="padding:0 0 12px 0;">
                        <a href="{escape(contact_no_url)}" style="display:inline-block;padding:12px 18px;border-radius:999px;background:#ffffff;color:#173a2a;text-decoration:none;font-size:14px;font-weight:700;">Je ne vais pas le contacter</a>
                      </td>
                    </tr>
                  </table>
                  <p style="margin:4px 0 0;font-size:14px;line-height:1.7;color:#e6efe9;">
                    Coordonnées directes du candidat : {escape(str(candidat.get('email') or 'Non renseigné'))}
                    {' - ' if candidate_phone and candidate_phone != 'Non renseigné' else ''}
                    {escape(candidate_phone) if candidate_phone and candidate_phone != 'Non renseigné' else ''}
                  </p>
                </div>
              </td>
            </tr>
            <tr>
              <td style="padding:24px 32px 32px;">
                <p style="margin:0;font-size:15px;line-height:1.7;">
                  Cordialement,<br>
                  <strong>{escape(signature_name)}</strong><br>
                  EURES beta
                </p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
""".strip()
    return recipient, subject, body_text, body_html


def send_brevo_transactional_email(to_email: str, subject: str, text_body: str, html_body: str):
    """Send one transactional email via Brevo."""
    brevo = get_brevo_config()
    if not brevo['api_key'] or not brevo['from_email']:
        raise RuntimeError('Brevo configuration is incomplete.')

    payload = {
        'sender': {
            'email': brevo['from_email'],
            'name': brevo['from_name'],
        },
        'to': [{'email': to_email}],
        'subject': subject,
        'textContent': text_body,
        'htmlContent': html_body,
    }
    resp = requests.post(
        'https://api.brevo.com/v3/smtp/email',
        headers={
            'accept': 'application/json',
            'content-type': 'application/json',
            'api-key': brevo['api_key'],
        },
        json=payload,
        timeout=30,
    )
    if resp.status_code not in {200, 201, 202}:
        raise RuntimeError(f'Brevo send failed: HTTP {resp.status_code} - {resp.text}')
    return _parse_response_json_safe(resp)


def list_eures_admin_matchings(status: str = 'all') -> list[dict]:
    """Return matchings joined with candidate/employer info for the admin UI."""
    config = get_eures_matching_config()
    candidate_config = get_form_config('eures-beta', 'candidate')
    employer_config = get_form_config('eures-beta', 'employer')
    if not config or not candidate_config or not employer_config:
        raise RuntimeError('EURES beta Grist configuration is incomplete.')

    headers = _eures_admin_headers(config)
    candidate_headers = _eures_admin_headers(candidate_config)
    employer_headers = _eures_admin_headers(employer_config)

    matchings = fetch_table_records(config['doc_id'], EURES_MATCHINGS_TABLE, headers)
    candidats = fetch_table_records(candidate_config['doc_id'], EURES_CANDIDATS_TABLE, candidate_headers)
    besoins = fetch_table_records(employer_config['doc_id'], EURES_BESOINS_TABLE, employer_headers)

    candidats_by_id = {
        str((rec.get('fields') or {}).get('id_tally') or (rec.get('fields') or {}).get('uuid') or ''): rec.get('fields', {})
        for rec in candidats if isinstance(rec, dict)
    }
    besoins_by_id = {
        str((rec.get('fields') or {}).get('id_tally') or (rec.get('fields') or {}).get('uuid') or ''): rec.get('fields', {})
        for rec in besoins if isinstance(rec, dict)
    }

    rows = []
    for rec in matchings:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        if not fields:
            continue
        scoring_status = str(fields.get('statut') or '').strip().lower()
        admin_status = _eures_admin_status(fields)
        if scoring_status not in {'a_valider', 'auto_envoyable'} and admin_status == 'pending':
            continue
        if status in {'pending', 'accepted', 'refused'} and admin_status != status:
            continue

        besoin_id = str(fields.get('besoin_id') or '')
        candidat_id = str(fields.get('candidat_id') or '')
        candidat = candidats_by_id.get(candidat_id, {})
        besoin = besoins_by_id.get(besoin_id, {})

        rows.append({
            'record_id': rec.get('id'),
            'besoin_id': besoin_id,
            'candidat_id': candidat_id,
            'score': _safe_int(fields.get('score')),
            'scoring_status': scoring_status,
            'admin_status': admin_status,
            'workflow_status': _matching_workflow_status(fields),
            'score_metier': _safe_int(fields.get('score_metier')),
            'score_langues': _safe_int(fields.get('score_langues')),
            'score_mobilite': _safe_int(fields.get('score_mobilite')),
            'score_disponibilite': _safe_int(fields.get('score_disponibilite')),
            'score_salaire': _safe_int(fields.get('score_salaire')),
            'raisons': _split_matching_text(fields.get('raisons')),
            'points_faibles': _split_matching_text(fields.get('points_faibles')),
            'date_calcul': fields.get('date_calcul', ''),
            'admin_decision_at': fields.get('admin_decision_at', ''),
            'admin_decision_by': fields.get('admin_decision_by', ''),
            'admin_decision_note': fields.get('admin_decision_note', ''),
            'workflow_status_updated_at': fields.get('workflow_status_updated_at', ''),
            'workflow_status_updated_by': fields.get('workflow_status_updated_by', ''),
            'sent_to_employer_at': fields.get('sent_to_employer_at', ''),
            'sent_to_employer_by': fields.get('sent_to_employer_by', ''),
            'employer_response': fields.get('employer_response', ''),
            'employer_response_at': fields.get('employer_response_at', ''),
            'mise_en_relation_at': fields.get('mise_en_relation_at', ''),
            'mise_en_relation_by': fields.get('mise_en_relation_by', ''),
            'embauche_confirmee_at': fields.get('embauche_confirmee_at', ''),
            'embauche_confirmee_by': fields.get('embauche_confirmee_by', ''),
            'candidat': {
                'nom': candidat.get('nom', ''),
                'email': candidat.get('email', ''),
                'telephone': candidat.get('telephone', ''),
                'ville': candidat.get('ville', ''),
                'pays': candidat.get('pays', ''),
                'metier': candidat.get('metier', ''),
                'langues': candidat.get('langues', ''),
                'mobilite': candidat.get('mobilite', ''),
                'disponibilite': candidat.get('disponibilite', ''),
                'competences': candidat.get('competences', ''),
            },
            'employeur': {
                'employeur': besoin.get('employeur', ''),
                'contact': besoin.get('contact', ''),
                'email': _coalesce_row_value(besoin, 'email', 'mail', 'contact_email', 'email_contact'),
                'pays': besoin.get('pays', ''),
                'poste': besoin.get('poste', ''),
                'langues_requises': besoin.get('langues_requises', ''),
                'date_debut': besoin.get('date_debut', ''),
                'competences_clefs': besoin.get('competences_clefs', ''),
                'contraintes_travail': besoin.get('contraintes_travail', ''),
            },
        })

    rows.sort(key=lambda row: (row['admin_status'] != 'pending', -row['score'], -(int(row['record_id'] or 0))))
    return rows


def normalize_finess(value) -> str:
    """Normalize FINESS value for duplicate checks."""
    raw = str(value or '').strip()
    digits = ''.join(ch for ch in raw if ch.isdigit())
    if digits and digits == raw:
        return digits.zfill(9) if len(digits) == 8 else digits
    return raw.upper()

def normalize_email(value) -> str:
    """Normalize email value for lookup."""
    return str(value or '').strip().lower()


def extract_finess_values(fields: dict) -> set:
    """Extract FINESS values from standard and JSON fields."""
    values = set()
    if not isinstance(fields, dict):
        return values

    main = normalize_finess(fields.get('finess_main'))
    if main:
        values.add(main)

    raw_json = fields.get('finess_json')
    parsed = []
    if isinstance(raw_json, str) and raw_json.strip():
        try:
            parsed = json.loads(raw_json)
        except Exception:
            parsed = []
    elif isinstance(raw_json, list):
        parsed = raw_json

    if isinstance(parsed, list):
        for item in parsed:
            v = normalize_finess(item)
            if v:
                values.add(v)

    return values


def find_duplicate_finess(config: dict, current_uuid: str, finess_values: set, headers: dict):
    """Find FINESS values already used by another record in the same table."""
    if not finess_values:
        return set()

    records = fetch_all_records(config, headers)
    duplicates = set()
    cur = normalize_finess(current_uuid)

    for rec in records:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        rec_uuid = normalize_finess(fields.get('uuid'))
        if cur and rec_uuid == cur:
            continue
        rec_finess = extract_finess_values(fields)
        duplicates.update(finess_values.intersection(rec_finess))

    return duplicates


@app.route('/api/forms/<form_id>/record', methods=['GET'])
def get_record(form_id: str):
    """Fetch a record by UUID or table-specific identifier."""
    config = get_form_config(form_id, request.args.get('flow_role'))
    if not config:
        return jsonify({'error': f'Unknown form: {form_id}'}), 404

    uuid = request.args.get('uuid')
    if not uuid:
        return jsonify({'error': 'UUID parameter required'}), 400

    # Build Grist API URL with filter
    filter_param = f'{{"uuid":["{uuid}"]}}'
    url = f"{GRIST_BASE_URL}/api/docs/{config['doc_id']}/tables/{config['table_id']}/records"

    headers = {'Accept': 'application/json'}
    # API key optional for read if doc is public
    if config['api_key']:
        headers['Authorization'] = f"Bearer {config['api_key']}"

    try:
        allowed_columns = get_table_columns(config, headers)
        record_key = resolve_record_key(allowed_columns)
        if not record_key:
            return jsonify({'error': 'Target table has no supported record key (uuid or id_tally).'}), 500
        resp = requests.get(
            url,
            params={'filter': json.dumps({record_key: [uuid]})},
            headers=headers,
        )
        return jsonify(resp.json()), resp.status_code
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/record', methods=['POST'])
def save_record(form_id: str):
    """Create or update a record."""
    data = request.get_json()
    if not data or 'fields' not in data:
        return jsonify({'error': 'Invalid request body'}), 400

    fields = data['fields']
    if not isinstance(fields, dict):
        return jsonify({'error': 'Invalid fields payload'}), 400
    config = get_form_config(form_id, fields.get('flow_role'))
    if not config:
        return jsonify({'error': f'Unknown form: {form_id}'}), 404

    if not config['api_key']:
        return jsonify({'error': 'API key not configured for this form'}), 500
    uuid = fields.get('uuid') or fields.get('id_tally')
    if not uuid:
        return jsonify({'error': 'UUID or id_tally required in fields'}), 400

    base_url = f"{GRIST_BASE_URL}/api/docs/{config['doc_id']}/tables/{config['table_id']}"
    headers = {
        'Authorization': f"Bearer {config['api_key']}",
        'Content-Type': 'application/json',
        'Accept': 'application/json',
    }

    # Keep only fields that exist in target table (prod/local may differ).
    try:
        allowed_columns = get_table_columns(config, headers)
        filtered_fields = {k: v for k, v in fields.items() if str(k) in allowed_columns}
        record_key = resolve_record_key(allowed_columns)
    except Exception as e:
        return jsonify({'error': f'Failed to fetch Grist table columns: {e}'}), 500

    if not record_key:
        return jsonify({'error': "Table is missing a supported unique key column ('uuid' or 'id_tally')"}), 500

    if record_key not in filtered_fields:
        filtered_fields[record_key] = uuid

    # Check if record exists
    try:
        existing_record, check_resp = fetch_record_by_field(base_url, record_key, filtered_fields[record_key], headers)
        if check_resp.status_code != 200:
            return jsonify(_parse_response_json_safe(check_resp)), check_resp.status_code
        record_id = existing_record.get('id') if isinstance(existing_record, dict) else None
    except Exception as e:
        return jsonify({'error': f'Failed to check existing record: {e}'}), 500

    # Enforce uniqueness of FINESS across records (excluding current UUID)
    try:
        incoming_finess = extract_finess_values(filtered_fields)
        duplicates = find_duplicate_finess(config, uuid, incoming_finess, headers)
        if duplicates:
            duplicates_sorted = sorted(duplicates)
            return jsonify({
                'error': 'Un ou plusieurs numéros FINESS sont déjà utilisés par un autre questionnaire.',
                'duplicates': duplicates_sorted,
            }), 409
    except Exception as e:
        return jsonify({'error': f'Failed to validate FINESS uniqueness: {e}'}), 500

    # Create or update, then re-read the UUID so callers never get a false success.
    try:
        action = 'updated' if record_id else 'created'
        if record_id:
            # Update existing
            payload = {'records': [{'id': record_id, 'fields': filtered_fields}]}
            resp = write_grist_records('PATCH', f"{base_url}/records", payload, headers)
        else:
            # Create new
            payload = {'records': [{'fields': filtered_fields}]}
            resp = write_grist_records('POST', f"{base_url}/records", payload, headers)

        redirect_error = _grist_write_redirect_response(resp)
        if redirect_error:
            return redirect_error

        if resp.status_code != 200:
            return jsonify(_parse_response_json_safe(resp)), resp.status_code

        saved_record, verify_resp = fetch_record_by_field(base_url, record_key, filtered_fields[record_key], headers)
        if verify_resp.status_code != 200:
            return jsonify({
                'error': f'Grist write returned success, but verification failed with HTTP {verify_resp.status_code}.',
            }), 502
        if not isinstance(saved_record, dict):
            return jsonify({
                'error': 'Grist write returned success, but the questionnaire was not found after saving.',
            }), 502

        saved_fields = saved_record.get('fields', {}) if isinstance(saved_record.get('fields'), dict) else {}
        if str(saved_fields.get(record_key) or '') != str(filtered_fields[record_key]):
            return jsonify({
                'error': f'Grist write returned success, but the saved questionnaire {record_key} did not match.',
            }), 502

        matching_result = None
        invitation_linking = None
        if form_id == 'eures-beta':
            if action == 'created':
                matching_result = run_eures_matching_for_saved_record(
                    form_id=form_id,
                    role=str(fields.get('flow_role') or ''),
                    saved_record=saved_record,
                    config=config,
                    headers=headers,
                )
            else:
                matching_result = {
                    'processed': False,
                    'reason': 'existing_record_update',
                    'role': str(fields.get('flow_role') or ''),
                }
            invitation_linking = link_eures_invitation_after_save(
                role=str(fields.get('flow_role') or ''),
                request_fields=fields,
                saved_record=saved_record,
                matching_result=matching_result,
            )

        return jsonify({
            'ok': True,
            'action': action,
            'uuid': uuid,
            'record_key': record_key,
            'record_id': saved_record.get('id'),
            'matching': matching_result,
            'invitation_linking': invitation_linking,
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/export-readable-xlsx', methods=['POST'])
def export_readable_xlsx(form_id: str):
    """Generate a human-readable Excel export from a form payload without changing Grist storage."""
    data = request.get_json()
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid request body'}), 400
    try:
        xlsx = build_readable_xlsx(data)
    except Exception as e:
        return jsonify({'error': f'Failed to build readable Excel export: {e}'}), 500

    uuid = str(data.get('uuid') or (data.get('fields') or {}).get('uuid') or 'saisie').strip() or 'saisie'
    filename = f'fagerh_saisie_lisible_{uuid}.xlsx'
    return send_file(
        xlsx,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@app.route('/api/forms/<form_id>/check-finess', methods=['POST'])
def check_finess(form_id: str):
    """Check whether FINESS values already exist in another questionnaire."""
    config = get_form_config(form_id)
    if not config:
        return jsonify({'error': f'Unknown form: {form_id}'}), 404

    data = request.get_json() or {}
    uuid = normalize_finess(data.get('uuid'))
    raw_finess = data.get('finess') or []

    if not isinstance(raw_finess, list):
        return jsonify({'error': 'Invalid request body: finess must be a list'}), 400

    finess_values = {normalize_finess(v) for v in raw_finess if normalize_finess(v)}
    headers = {'Accept': 'application/json'}
    if config['api_key']:
        headers['Authorization'] = f"Bearer {config['api_key']}"

    try:
        duplicates = sorted(find_duplicate_finess(config, uuid, finess_values, headers))
        return jsonify({
            'ok': True,
            'duplicates': duplicates,
            'has_duplicates': len(duplicates) > 0,
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/recover-by-email', methods=['POST'])
def recover_by_email(form_id: str):
    """Recover a questionnaire UUID from validation email (+ optional FINESS)."""
    config = get_form_config(form_id)
    if not config:
        return jsonify({'error': f'Unknown form: {form_id}'}), 404

    data = request.get_json() or {}
    email = normalize_email(data.get('email'))
    finess = normalize_finess(data.get('finess'))
    if not email:
        return jsonify({'error': 'Email required'}), 400
    if finess and not finess.isdigit():
        return jsonify({'error': 'FINESS invalide'}), 400

    url = f"{GRIST_BASE_URL}/api/docs/{config['doc_id']}/tables/{config['table_id']}/records"
    headers = {'Accept': 'application/json'}
    if config['api_key']:
        headers['Authorization'] = f"Bearer {config['api_key']}"

    try:
        # 1) Fast path: exact filter on the email column.
        filter_param = json.dumps({"validateur_email": [email]})
        resp = requests.get(url, params={'filter': filter_param, 'limit': 5000}, headers=headers)
        if resp.status_code != 200:
            return jsonify(resp.json()), resp.status_code
        payload = resp.json()
        records = payload.get('records', [])

        matches = []
        for rec in records:
            fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
            if not fields.get('uuid'):
                continue
            if finess:
                rec_finess = extract_finess_values(fields)
                if finess not in rec_finess:
                    continue
            matches.append(rec)

        # 2) Fallback: scan and compare case-insensitively (older rows / casing differences).
        if not matches:
            scan_resp = requests.get(url, params={'limit': 5000}, headers=headers)
            if scan_resp.status_code != 200:
                return jsonify(scan_resp.json()), scan_resp.status_code
            scan_payload = scan_resp.json()
            for rec in scan_payload.get('records', []):
                fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
                if normalize_email(fields.get('validateur_email')) != email:
                    continue
                if not fields.get('uuid'):
                    continue
                if finess:
                    rec_finess = extract_finess_values(fields)
                    if finess not in rec_finess:
                        continue
                matches.append(rec)

        if not matches:
            if finess:
                return jsonify({'error': 'Aucun questionnaire trouvé pour ce couple email + FINESS.'}), 404
            return jsonify({'error': 'Aucun questionnaire trouvé pour cet email.'}), 404

        chosen = max(matches, key=lambda r: int(r.get('id', 0) or 0))
        chosen_fields = chosen.get('fields', {}) if isinstance(chosen, dict) else {}
        return jsonify({
            'ok': True,
            'uuid': chosen_fields.get('uuid'),
            'count': len(matches),
            'match_on': 'email+finess' if finess else 'email',
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/admin/overview', methods=['GET'])
@admin_required
def admin_overview(form_id: str):
    """Admin dashboard data: counts + questionnaire list."""
    config = get_form_config(form_id)
    if not config:
        return jsonify({'error': f'Unknown form: {form_id}'}), 404

    headers = {'Accept': 'application/json'}
    if config['api_key']:
        headers['Authorization'] = f"Bearer {config['api_key']}"

    search = str(request.args.get('search', '') or '').strip().lower()
    status = str(request.args.get('status', 'all') or 'all').strip().lower()

    try:
        records = fetch_all_records(config, headers)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    rows = []
    for rec in records:
        fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
        if not fields:
            continue
        row = {
            'record_id': rec.get('id'),
            'uuid': fields.get('uuid', ''),
            'es_nom': fields.get('es_nom', ''),
            'departement': fields.get('es_departement', ''),
            'finess_main': fields.get('finess_main', ''),
            'validateur_nom': fields.get('validateur_nom', ''),
            'validateur_prenom': fields.get('validateur_prenom', ''),
            'validateur_email': fields.get('validateur_email', ''),
            'saisie_terminee': _as_bool(fields.get('saisie_terminee')),
        }
        row['quick_progress'] = compute_quick_step_progress(fields)
        rows.append(row)

    total = len(rows)
    termines = sum(1 for r in rows if r['saisie_terminee'])
    en_cours = total - termines

    # Newest first by record id.
    rows.sort(key=lambda r: int(r['record_id'] or 0), reverse=True)

    if status in {'en_cours', 'termines'}:
        want_done = status == 'termines'
        rows = [r for r in rows if r['saisie_terminee'] is want_done]

    if search:
        def _matches(r):
            haystack = ' '.join([
                str(r.get('es_nom', '')),
                str(r.get('departement', '')),
                str(r.get('finess_main', '')),
                str(r.get('validateur_nom', '')),
                str(r.get('validateur_prenom', '')),
                str(r.get('validateur_email', '')),
                str(r.get('uuid', '')),
            ]).lower()
            return search in haystack
        rows = [r for r in rows if _matches(r)]

    return jsonify({
        'ok': True,
        'form_id': form_id,
        'stats': {
            'total_questionnaires': total,
            'en_cours': en_cours,
            'termines': termines,
        },
        'rows': rows,
    }), 200


@app.route('/api/forms/<form_id>/admin/invitations', methods=['GET'])
@admin_required
def admin_eures_invitations(form_id: str):
    """Admin API: list invitations for EURES beta."""
    if form_id != 'eures-beta':
        return jsonify({'error': f'Unknown invitation admin form: {form_id}'}), 404

    status = str(request.args.get('status', 'all') or 'all').strip().lower()
    role = _normalize_eures_invitation_role(request.args.get('role', '')) if request.args.get('role') else ''
    search = str(request.args.get('search', '') or '').strip().lower()

    try:
        rows = list_eures_invitations()
        if status != 'all':
            rows = [row for row in rows if str(row.get('invitation_status') or '') == status]
        if role:
            rows = [row for row in rows if str(row.get('role') or '') == role]
        if search:
            rows = [
                row for row in rows
                if search in ' '.join([
                    str(row.get('email') or ''),
                    str(row.get('first_name') or ''),
                    str(row.get('last_name') or ''),
                    str(row.get('company_name') or ''),
                    str(row.get('external_ref') or ''),
                ]).lower()
            ]
        stats = Counter(row.get('invitation_status') or 'invitation_a_envoyer' for row in rows)
        return jsonify({
            'ok': True,
            'form_id': form_id,
            'stats': {
                'total': len(rows),
                'invitation_a_envoyer': stats.get('invitation_a_envoyer', 0),
                'invitation_envoyee': stats.get('invitation_envoyee', 0),
                'questionnaire_recu': stats.get('questionnaire_recu', 0),
                'rapprochee': stats.get('rapprochee', 0),
                'erreur_envoi': stats.get('erreur_envoi', 0),
            },
            'rows': rows,
        }), 200
    except Exception as e:
        app.logger.exception('EURES invitations list failed')
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/admin/invitations/import', methods=['POST'])
@admin_required
def admin_eures_invitations_import(form_id: str):
    """Admin API: import invitation rows from JSON or CSV payload."""
    if form_id != 'eures-beta':
        return jsonify({'error': f'Unknown invitation import form: {form_id}'}), 404

    data = request.get_json() or {}
    rows = _parse_eures_invitation_rows(data)
    if not rows:
        return jsonify({'error': 'No invitation rows provided'}), 400

    auth = request.authorization
    actor = auth.username if auth and auth.username else 'admin'
    try:
        result = upsert_eures_invitation_rows(rows, actor=actor)
        return jsonify(result), 200
    except Exception as e:
        app.logger.exception('EURES invitations import failed')
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/admin/invitations/<int:record_id>', methods=['PATCH'])
@admin_required
def admin_eures_invitation_update(form_id: str, record_id: int):
    """Admin API: patch one invitation row."""
    if form_id != 'eures-beta':
        return jsonify({'error': f'Unknown invitation update form: {form_id}'}), 404

    data = request.get_json() or {}
    update_fields = {}
    role = data.get('role')
    if role is not None:
        normalized_role = _normalize_eures_invitation_role(role)
        if normalized_role not in EURES_INVITATION_ALLOWED_ROLES:
            return jsonify({'error': 'Invalid invitation role'}), 400
        update_fields['role'] = normalized_role

    email = data.get('email')
    if email is not None:
        normalized_email = normalize_email(email)
        if not normalized_email:
            return jsonify({'error': 'Invalid invitation email'}), 400
        update_fields['email'] = normalized_email

    status = data.get('invitation_status')
    if status is not None:
        normalized_status = _normalize_eures_invitation_status(status)
        if normalized_status not in EURES_INVITATION_ALLOWED_STATUSES:
            return jsonify({'error': 'Invalid invitation status'}), 400
        auth = request.authorization
        actor = auth.username if auth and auth.username else 'admin'
        update_fields['invitation_status'] = normalized_status
        update_fields['invitation_status_updated_at'] = _now_iso_utc()
        update_fields['invitation_status_updated_by'] = actor

    for field_name in {
        'first_name',
        'last_name',
        'company_name',
        'language',
        'source',
        'external_ref',
        'notes',
        'answered_at',
        'linked_form_role',
        'linked_record_id',
        'linked_record_key',
        'matching_status',
    }:
        if field_name in data:
            update_fields[field_name] = str(data.get(field_name) or '').strip()

    if not update_fields:
        return jsonify({'error': 'No supported invitation fields provided'}), 400

    config = get_eures_invitations_config()
    if not config:
        return jsonify({'error': 'EURES invitations configuration is incomplete.'}), 500

    headers = _eures_admin_headers(config)
    try:
        ensure_table_columns(config, set(update_fields.keys()) & EURES_INVITATION_FIELDS, headers)
        allowed_columns = get_table_columns(config, headers)
        filtered_fields = {k: v for k, v in update_fields.items() if k in allowed_columns}
        base_url = f"{GRIST_BASE_URL}/api/docs/{config['doc_id']}/tables/{config['table_id']}/records"
        resp = write_grist_records('PATCH', base_url, {'records': [{'id': record_id, 'fields': filtered_fields}]}, headers)
        if resp.status_code != 200:
            raise RuntimeError(f'Failed to update invitation: HTTP {resp.status_code} - {resp.text}')
        updated = fetch_record_by_id(config['doc_id'], config['table_id'], record_id, headers)
        return jsonify({
            'ok': True,
            'record_id': record_id,
            'fields': updated.get('fields', {}) if isinstance(updated, dict) else {},
        }), 200
    except Exception as e:
        app.logger.exception('EURES invitation update failed')
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/admin/invitations/send', methods=['POST'])
@admin_required
def admin_eures_invitations_send(form_id: str):
    """Admin API: send pending invitations through Brevo."""
    if form_id != 'eures-beta':
        return jsonify({'error': f'Unknown invitation send form: {form_id}'}), 404

    data = request.get_json() or {}
    requested_ids = data.get('record_ids') or []
    if requested_ids and not isinstance(requested_ids, list):
        return jsonify({'error': 'record_ids must be a list'}), 400

    config = get_eures_invitations_config()
    if not config:
        return jsonify({'error': 'EURES invitations configuration is incomplete.'}), 500

    headers = _eures_admin_headers(config)
    auth = request.authorization
    actor = auth.username if auth and auth.username else 'admin'
    try:
        records = fetch_table_records(config['doc_id'], config['table_id'], headers)
        if requested_ids:
            requested_ids_set = {int(value) for value in requested_ids}
            target_records = [rec for rec in records if int(rec.get('id') or 0) in requested_ids_set]
        else:
            target_records = []
            for rec in records:
                fields = rec.get('fields', {}) if isinstance(rec, dict) else {}
                if str(fields.get('invitation_status') or '').strip().lower() == 'invitation_a_envoyer':
                    target_records.append(rec)

        sent = []
        skipped = []
        errors = []
        for rec in target_records:
            record_id = int(rec.get('id') or 0)
            fields = rec.get('fields', {}) if isinstance(rec.get('fields'), dict) else {}
            current_status = str(fields.get('invitation_status') or '').strip().lower()
            if current_status not in EURES_INVITATION_SENDABLE_STATUSES:
                skipped.append({
                    'record_id': record_id,
                    'email': fields.get('email', ''),
                    'reason': f'status_not_sendable:{current_status or "unknown"}',
                })
                continue
            try:
                recipient, subject, text_body, html_body, invite_token, invite_link = build_brevo_invitation_email(fields)
                brevo_result = send_brevo_transactional_email(recipient, subject, text_body, html_body)
                update_eures_invitation_record_by_id(record_id, {
                    'invite_token': invite_token,
                    'invite_link': invite_link,
                    'invitation_status': 'invitation_envoyee',
                    'invitation_status_updated_at': _now_iso_utc(),
                    'invitation_status_updated_by': actor,
                    'sent_at': _now_iso_utc(),
                    'sent_by': actor,
                    'brevo_message_id': str((brevo_result or {}).get('messageId') or ''),
                }, headers=headers)
                sent.append({
                    'record_id': record_id,
                    'email': recipient,
                    'brevo_message_id': str((brevo_result or {}).get('messageId') or ''),
                })
            except Exception as e:
                error_message = str(e)
                errors.append({
                    'record_id': record_id,
                    'email': fields.get('email', ''),
                    'error': error_message,
                })
                try:
                    update_eures_invitation_record_by_id(record_id, {
                        'invitation_status': 'erreur_envoi',
                        'invitation_status_updated_at': _now_iso_utc(),
                        'invitation_status_updated_by': actor,
                        'notes': error_message[:500],
                    }, headers=headers)
                except Exception:
                    app.logger.exception('EURES invitation error status update failed')

        return jsonify({
            'ok': True,
            'requested': len(target_records),
            'sent': sent,
            'skipped': skipped,
            'errors': errors,
        }), 200
    except Exception as e:
        app.logger.exception('EURES invitations send failed')
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/admin/matchings', methods=['GET'])
@admin_required
def admin_eures_matchings(form_id: str):
    """Admin API: list matchings for EURES beta review."""
    if form_id != 'eures-beta':
        return jsonify({'error': f'Unknown admin matching form: {form_id}'}), 404

    status = str(request.args.get('status', 'all') or 'all').strip().lower()
    if status not in {'all', 'pending', 'accepted', 'refused'}:
        return jsonify({'error': 'Invalid status filter'}), 400

    try:
        rows = list_eures_admin_matchings(status=status)
        stats = Counter(row['admin_status'] for row in rows)
        return jsonify({
            'ok': True,
            'form_id': form_id,
            'stats': {
                'total': len(rows),
                'pending': stats.get('pending', 0),
                'accepted': stats.get('accepted', 0),
                'refused': stats.get('refused', 0),
            },
            'rows': rows,
        }), 200
    except Exception as e:
        app.logger.exception('EURES admin matchings list failed')
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/admin/brevo-health', methods=['GET'])
@admin_required
def admin_eures_brevo_health(form_id: str):
    """Admin API: return the current Brevo configuration and API reachability state."""
    if form_id != 'eures-beta':
        return jsonify({'error': f'Unknown admin form: {form_id}'}), 404

    check_api = str(request.args.get('check') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
    health = get_brevo_health(check_api=check_api)
    status_code = 200 if health['configured'] and (not check_api or health['api_ok'] is True) else 503
    return jsonify({
        'ok': status_code == 200,
        'brevo': health,
    }), status_code


@app.route('/api/forms/<form_id>/admin/matchings/<int:record_id>/decision', methods=['POST'])
@admin_required
def admin_eures_matching_decision(form_id: str, record_id: int):
    """Admin API: accept or refuse one matching."""
    if form_id != 'eures-beta':
        return jsonify({'error': f'Unknown admin matching form: {form_id}'}), 404

    data = request.get_json() or {}
    decision = str(data.get('decision') or '').strip().lower()
    note = str(data.get('note') or '').strip()
    if decision not in {'accepted', 'refused'}:
        return jsonify({'error': 'Decision must be accepted or refused'}), 400

    config = get_eures_matching_config()
    if not config:
        return jsonify({'error': 'EURES beta matching configuration is incomplete.'}), 500

    headers = _eures_admin_headers(config)
    try:
        existing = fetch_record_by_id(config['doc_id'], EURES_MATCHINGS_TABLE, record_id, headers)
        if not existing:
            return jsonify({'error': 'Matching not found'}), 404
        existing_fields = existing.get('fields', {}) if isinstance(existing.get('fields'), dict) else {}

        if decision == 'accepted':
            ensure_brevo_ready(check_api=True)

        auth = request.authorization
        decided_by = auth.username if auth and auth.username else 'admin'
        decision_fields = {
            'admin_status': decision,
            'admin_decision_at': _now_iso_utc(),
            'admin_decision_by': decided_by,
            'admin_decision_note': note,
        }
        if decision == 'accepted':
            decision_fields.update(_matching_workflow_update_fields('valide_admin', decided_by))
        else:
            decision_fields.update(_matching_workflow_update_fields('refuse_admin', decided_by))
        update_matching_record_by_id(config['doc_id'], record_id, decision_fields, headers)

        app.logger.info(
            'EURES matching decision saved',
            extra={
                'form_id': form_id,
                'record_id': record_id,
                'decision': decision,
                'decided_by': decided_by,
            },
        )

        updated = fetch_record_by_id(config['doc_id'], EURES_MATCHINGS_TABLE, record_id, headers)
        updated_fields = updated.get('fields', {}) if isinstance(updated, dict) else {}
        brevo_result = None
        if decision == 'accepted':
            matching_rows = list_eures_admin_matchings(status='all')
            matching_row = next((row for row in matching_rows if int(row.get('record_id') or 0) == record_id), None)
            if not matching_row:
                raise RuntimeError('Accepted matching could not be reloaded for email delivery.')
            to_email, subject, text_body, html_body = build_brevo_matching_email(matching_row)
            brevo_result = send_brevo_transactional_email(to_email, subject, text_body, html_body)
            update_matching_record_by_id(
                config['doc_id'],
                record_id,
                _matching_workflow_update_fields('envoye_employeur', decided_by),
                headers,
            )
            app.logger.info(
                'EURES accepted matching email sent',
                extra={
                    'form_id': form_id,
                    'record_id': record_id,
                    'to_email': to_email,
                },
            )
        return jsonify({
            'ok': True,
            'record_id': record_id,
            'admin_status': updated_fields.get('admin_status', decision),
            'admin_decision_at': updated_fields.get('admin_decision_at', decision_fields['admin_decision_at']),
            'admin_decision_by': updated_fields.get('admin_decision_by', decided_by),
            'admin_decision_note': updated_fields.get('admin_decision_note', note),
            'email_sent': decision == 'accepted',
            'email_result': brevo_result,
        }), 200
    except Exception as e:
        app.logger.exception('EURES matching decision update failed')
        return jsonify({'error': str(e)}), 500


@app.route('/api/forms/<form_id>/admin/matchings/<int:record_id>/workflow', methods=['POST'])
@admin_required
def admin_eures_matching_workflow(form_id: str, record_id: int):
    """Admin API: advance one matching in the business workflow."""
    if form_id != 'eures-beta':
        return jsonify({'error': f'Unknown admin matching form: {form_id}'}), 404

    data = request.get_json() or {}
    target_status = str(data.get('workflow_status') or '').strip().lower()
    note = str(data.get('note') or '').strip()
    allowed_targets = {'mise_en_relation_faite', 'embauche_confirmee', 'envoye_employeur'}
    if target_status not in allowed_targets:
        return jsonify({'error': 'Unsupported workflow status'}), 400

    config = get_eures_matching_config()
    if not config:
        return jsonify({'error': 'EURES beta matching configuration is incomplete.'}), 500

    headers = _eures_admin_headers(config)
    try:
        existing = fetch_record_by_id(config['doc_id'], EURES_MATCHINGS_TABLE, record_id, headers)
        if not existing:
            return jsonify({'error': 'Matching not found'}), 404
        existing_fields = existing.get('fields', {}) if isinstance(existing.get('fields'), dict) else {}
        current_status = _matching_workflow_status(existing_fields)
        if not _eures_workflow_transition_allowed(current_status, target_status):
            return jsonify({
                'error': f'Transition impossible: {current_status} -> {target_status}',
            }), 400

        auth = request.authorization
        actor = auth.username if auth and auth.username else 'admin'
        update_fields = _matching_workflow_update_fields(target_status, actor)
        if note:
            previous_note = str(existing_fields.get('admin_decision_note') or '').strip()
            update_fields['admin_decision_note'] = f'{previous_note}\n{note}'.strip() if previous_note else note
        update_matching_record_by_id(config['doc_id'], record_id, update_fields, headers)
        updated = fetch_record_by_id(config['doc_id'], EURES_MATCHINGS_TABLE, record_id, headers)
        updated_fields = updated.get('fields', {}) if isinstance(updated, dict) else {}
        return jsonify({
            'ok': True,
            'record_id': record_id,
            'workflow_status': _matching_workflow_status(updated_fields),
            'workflow_label': _matching_workflow_label(_matching_workflow_status(updated_fields)),
            'workflow_status_updated_at': updated_fields.get('workflow_status_updated_at', ''),
            'workflow_status_updated_by': updated_fields.get('workflow_status_updated_by', actor),
        }), 200
    except Exception as e:
        app.logger.exception('EURES matching workflow update failed')
        return jsonify({'error': str(e)}), 500


@app.route('/eures-beta/matching-feedback', methods=['GET'])
def eures_matching_feedback():
    """Record employer feedback from email CTA links."""
    token = str(request.args.get('token') or '').strip()
    if not token:
        return Response('Lien invalide : token manquant.', status=400, mimetype='text/plain')

    try:
        payload = get_eures_email_action_serializer().loads(token)
    except BadSignature:
        return Response('Lien invalide ou expiré.', status=400, mimetype='text/plain')
    except Exception as e:
        app.logger.exception('EURES feedback token decode failed')
        return Response(f'Erreur de lecture du lien : {e}', status=500, mimetype='text/plain')

    record_id = int(payload.get('record_id') or 0)
    response_code = str(payload.get('response') or '').strip().lower()
    if not record_id or response_code not in {'contact', 'not_contact'}:
        return Response('Lien invalide : données incomplètes.', status=400, mimetype='text/plain')

    config = get_eures_matching_config()
    if not config:
        return Response('Configuration EURES incomplète.', status=500, mimetype='text/plain')

    headers = _eures_admin_headers(config)
    response_label = 'je vais le contacter' if response_code == 'contact' else 'je ne vais pas le contacter'
    update_fields = _matching_workflow_update_fields(
        'accepte_employeur' if response_code == 'contact' else 'refuse_employeur',
        'employer_email_link',
    )

    try:
        update_matching_record_by_id(config['doc_id'], record_id, update_fields, headers)
        app.logger.info(
            'EURES employer response recorded',
            extra={
                'record_id': record_id,
                'response': response_code,
            },
        )
    except Exception as e:
        app.logger.exception('EURES employer response update failed')
        return Response(f'Erreur lors de l’enregistrement de votre réponse : {e}', status=500, mimetype='text/plain')

    html = f"""<!doctype html>
<html lang="fr">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>EURES beta - Réponse enregistrée</title>
  </head>
  <body style="margin:0;background:#f4efe6;font-family:Georgia,'Times New Roman',serif;color:#1f1f1f;">
    <div style="max-width:720px;margin:32px auto;padding:0 16px;">
      <div style="background:#fffdf9;border:1px solid #e7dcc7;border-radius:18px;overflow:hidden;">
        <div style="padding:28px 32px;background:linear-gradient(135deg,#103a2b 0%,#1f5a45 100%);color:#fff;">
          <div style="font-size:13px;letter-spacing:1.4px;text-transform:uppercase;opacity:0.82;">EURES beta</div>
          <h1 style="margin:10px 0 0;font-size:30px;line-height:1.2;">Réponse enregistrée</h1>
        </div>
        <div style="padding:28px 32px;">
          <p style="margin:0 0 16px;font-size:17px;line-height:1.7;">
            Merci. Votre réponse a bien été enregistrée :
            <strong>{escape(response_label)}</strong>.
          </p>
          <p style="margin:0;font-size:15px;line-height:1.7;">
            Si besoin, vous pouvez aussi répondre directement à l’email reçu.
          </p>
        </div>
      </div>
    </div>
  </body>
</html>"""
    return Response(html, status=200, mimetype='text/html')


@app.route('/api/forms/<form_id>/public-stats', methods=['GET'])
def public_stats(form_id: str):
    """Public aggregated stats page data."""
    if form_id != 'eures-beta':
        return jsonify({'error': f'Unknown public stats form: {form_id}'}), 404

    try:
        return jsonify(build_eures_public_stats()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/health')
def health():
    """Health check endpoint."""
    deep = str(request.args.get('deep') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
    payload = {'status': 'ok'}
    if deep:
        brevo = get_brevo_health(check_api=True)
        payload['brevo'] = brevo
        if not brevo['configured'] or brevo['api_ok'] is not True:
            payload['status'] = 'degraded'
            return jsonify(payload), 503
    return jsonify(payload)


@app.route('/')
def index():
    return redirect('/forms/fagerh/')


# Static file serving for forms and frontend
@app.route('/forms/<form_id>/')
@app.route('/forms/<form_id>')
def serve_form(form_id: str):
    """Serve form HTML."""
    return send_from_directory(FORMS_DIR / form_id, 'index.html')


@app.route('/forms/fagerh/questions-pdf')
def serve_fagerh_questions_pdf():
    """Serve the reference PDF containing FAGERH questions."""
    return send_from_directory(
        DOCS_DIR,
        'fagerh_questions_completes.pdf',
        as_attachment=True,
        download_name='fagerh_questions_fagerh.pdf',
    )


@app.route('/forms/<form_id>/<path:filename>')
def serve_form_file(form_id: str, filename: str):
    """Serve extra files from a form folder (e.g. UI prototypes)."""
    resolved = _resolve_form_path(form_id, filename)
    if not resolved:
        return jsonify({'error': 'File not found'}), 404
    return send_from_directory(FORMS_DIR / form_id, resolved)


@app.route('/admin/<form_id>/')
@app.route('/admin/<form_id>')
@admin_required
def serve_admin(form_id: str):
    """Serve admin dashboard HTML for a form."""
    return send_from_directory(FORMS_DIR / form_id, 'admin.html')


@app.route('/assets/<path:filename>')
def serve_assets(filename: str):
    """Serve static assets (JS, CSS)."""
    return send_from_directory(ASSETS_DIR, filename)

if fagerh_suivi_app is not None:
    app.wsgi_app = DispatcherMiddleware(
        app.wsgi_app,
        {
            '/forms/fagerh/saisie-quotidienne': fagerh_suivi_app,
        },
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5005, debug=False)
